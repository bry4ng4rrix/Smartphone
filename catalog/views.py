import io

import openpyxl
from django.db import transaction
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from users.permissions import IsGerantOrReadOnly, get_accessible_magasins, resolve_magasin_for_request

from . import services
from .models import Brand, Color, ProductCategory, ProductReference, ProductType, ProductVariant, StockMovement
from .serializers import (
    BrandSerializer,
    BulkPriceUpdateSerializer,
    ColorSerializer,
    ProductCategorySerializer,
    ProductReferenceAutocompleteSerializer,
    ProductReferenceSerializer,
    ProductTypeSerializer,
    ProductVariantSerializer,
    StockAdjustmentSerializer,
    StockMovementSerializer,
)


class ProductCategoryViewSet(viewsets.ModelViewSet):
    serializer_class = ProductCategorySerializer
    permission_classes = [IsGerantOrReadOnly]

    def get_queryset(self):
        qs = ProductCategory.objects.filter(magasin__in=get_accessible_magasins(self.request.user))
        magasin_id = self.request.query_params.get("magasin_id")
        if magasin_id:
            qs = qs.filter(magasin_id=magasin_id)
        return qs

    def perform_create(self, serializer):
        serializer.save(magasin=resolve_magasin_for_request(self.request))

    def destroy(self, request, *args, **kwargs):
        category = self.get_object()
        if category.types.exists():
            raise ValidationError(
                "Impossible de supprimer une catégorie qui a des sous-types — "
                "supprimez ou déplacez d'abord ses sous-types."
            )
        return super().destroy(request, *args, **kwargs)


class ProductTypeViewSet(viewsets.ModelViewSet):
    serializer_class = ProductTypeSerializer
    permission_classes = [IsGerantOrReadOnly]

    def get_queryset(self):
        qs = ProductType.objects.select_related("category").filter(
            category__magasin__in=get_accessible_magasins(self.request.user)
        )
        category_id = self.request.query_params.get("category")
        if category_id:
            qs = qs.filter(category_id=category_id)
        return qs

    def destroy(self, request, *args, **kwargs):
        type_obj = self.get_object()
        if type_obj.references.exists():
            raise ValidationError(
                "Impossible de supprimer un sous-type qui a des références produit — "
                "supprimez ou déplacez d'abord ses références."
            )
        return super().destroy(request, *args, **kwargs)


class BrandViewSet(viewsets.ModelViewSet):
    serializer_class = BrandSerializer
    permission_classes = [IsGerantOrReadOnly]

    def get_queryset(self):
        qs = Brand.objects.filter(magasin__in=get_accessible_magasins(self.request.user))
        magasin_id = self.request.query_params.get("magasin_id")
        if magasin_id:
            qs = qs.filter(magasin_id=magasin_id)
        return qs

    def perform_create(self, serializer):
        serializer.save(magasin=resolve_magasin_for_request(self.request))

    def destroy(self, request, *args, **kwargs):
        brand = self.get_object()
        if brand.references.exists():
            raise ValidationError(
                "Impossible de supprimer une marque qui a des références produit — "
                "supprimez ou déplacez d'abord ses références."
            )
        return super().destroy(request, *args, **kwargs)


class ColorViewSet(viewsets.ModelViewSet):
    serializer_class = ColorSerializer
    permission_classes = [IsGerantOrReadOnly]

    def get_queryset(self):
        qs = Color.objects.filter(magasin__in=get_accessible_magasins(self.request.user))
        magasin_id = self.request.query_params.get("magasin_id")
        if magasin_id:
            qs = qs.filter(magasin_id=magasin_id)
        return qs

    def perform_create(self, serializer):
        serializer.save(magasin=resolve_magasin_for_request(self.request))


class ProductReferenceViewSet(viewsets.ModelViewSet):
    serializer_class = ProductReferenceSerializer
    permission_classes = [IsGerantOrReadOnly]

    def get_queryset(self):
        qs = ProductReference.objects.select_related("type", "brand", "type__category").prefetch_related(
            "variants"
        ).filter(type__category__magasin__in=get_accessible_magasins(self.request.user))
        type_id = self.request.query_params.get("type")
        brand_id = self.request.query_params.get("brand")
        category_id = self.request.query_params.get("category")
        magasin_id = self.request.query_params.get("magasin_id")
        if type_id:
            qs = qs.filter(type_id=type_id)
        if brand_id:
            qs = qs.filter(brand_id=brand_id)
        if category_id:
            qs = qs.filter(type__category_id=category_id)
        if magasin_id:
            qs = qs.filter(type__category__magasin_id=magasin_id)
        return qs

    @action(detail=False, methods=["get"])
    def autocomplete(self, request):
        """GET /api/catalog/references/autocomplete/?q=...&type=<id>
        Recherche pour le formulaire Nouvelle commande (§6 Smartreadme.md)."""
        qs = self.get_queryset().filter(actif=True)
        q = request.query_params.get("q")
        if q:
            qs = qs.filter(reference_name__icontains=q)
        qs = qs[:20]
        return Response(ProductReferenceAutocompleteSerializer(qs, many=True).data)

    @action(detail=False, methods=["post"], url_path="bulk-update-price")
    def bulk_update_price(self, request):
        """POST /api/catalog/references/bulk-update-price/
        {type_id, prix_achat?, prix_vente?} — modifie prix_achat/prix_vente
        pour TOUTES les références d'un sous-type donné (ex: toutes les
        "Flip cover", quelle que soit la marque), en une seule action."""
        serializer = BulkPriceUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        qs = self.get_queryset().filter(type_id=data["type_id"])
        count = qs.count()
        if count == 0:
            return Response({"error": "Aucune référence pour ce sous-type."}, status=404)

        update_fields = {}
        if "prix_achat" in data:
            update_fields["prix_achat"] = data["prix_achat"]
        if "prix_vente" in data:
            update_fields["prix_vente"] = data["prix_vente"]
        qs.update(**update_fields)

        return Response({"updated": count})

    EXCEL_HEADERS = [
        "Catégorie", "Sous-type", "Marque", "Référence",
        "Couleur", "Prix achat", "Prix vente", "Stock actuel",
        "Seuil alerte", "Actif",
    ]

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        """GET /api/catalog/references/export-excel/ — une ligne par couleur
        (variante), pour édition hors-ligne puis réimport via import-excel/."""
        refs = self.get_queryset().order_by("type__category__nom", "type__nom", "brand__nom", "reference_name")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catalogue"
        ws.append(self.EXCEL_HEADERS)

        for ref in refs:
            variants = list(ref.variants.all())
            if not variants:
                ws.append([
                    ref.type.category.nom, ref.type.nom, ref.brand.nom, ref.reference_name,
                    "", float(ref.prix_achat or 0), float(ref.prix_vente or 0), "", "",
                    "Oui" if ref.actif else "Non",
                ])
            for v in variants:
                ws.append([
                    ref.type.category.nom, ref.type.nom, ref.brand.nom, ref.reference_name,
                    v.couleur, float(ref.prix_achat or 0), float(ref.prix_vente or 0),
                    v.stock_actuel, v.seuil_alerte, "Oui" if ref.actif else "Non",
                ])

        for i in range(1, len(self.EXCEL_HEADERS) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="catalogue.xlsx"'
        return response

    @action(detail=False, methods=["post"], url_path="import-excel", parser_classes=[MultiPartParser])
    def import_excel(self, request):
        """POST /api/catalog/references/import-excel/ (multipart, champ
        "file") — crée/actualise Catégorie → Sous-type → Marque → Référence →
        Couleur à partir d'un fichier au format export-excel/. Le stock ne
        change jamais directement (§10 Smartreadme.md) : un écart avec le
        stock actuel déclenche un mouvement ENTREE/SORTIE tracé."""
        uploaded = request.FILES.get("file")
        if not uploaded:
            return Response({"error": "Fichier requis (champ 'file')."}, status=400)

        try:
            wb = openpyxl.load_workbook(uploaded, data_only=True)
            ws = wb.active
        except Exception:
            return Response({"error": "Fichier Excel invalide."}, status=400)

        magasin = resolve_magasin_for_request(request)
        created_refs = 0
        updated_refs = 0
        created_variants = 0
        updated_variants = 0
        errors = []

        category_cache = {}
        type_cache = {}
        brand_cache = {}

        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(c in (None, "") for c in row):
                continue
            (categorie, sous_type, marque, reference_name, couleur,
             prix_achat, prix_vente, stock_actuel, seuil_alerte, actif) = (list(row) + [None] * 10)[:10]

            categorie = str(categorie or "").strip()
            sous_type = str(sous_type or "").strip()
            marque = str(marque or "").strip()
            reference_name = str(reference_name or "").strip()
            couleur = str(couleur or "").strip()

            if not (categorie and sous_type and marque and reference_name):
                errors.append(f"Ligne {row_index} : Catégorie/Sous-type/Marque/Référence manquant(e).")
                continue

            try:
                with transaction.atomic():
                    cat_key = (categorie,)
                    category = category_cache.get(cat_key)
                    if category is None:
                        category, _ = ProductCategory.objects.get_or_create(magasin=magasin, nom=categorie)
                        category_cache[cat_key] = category

                    type_key = (categorie, sous_type)
                    type_obj = type_cache.get(type_key)
                    if type_obj is None:
                        type_obj, _ = ProductType.objects.get_or_create(category=category, nom=sous_type)
                        type_cache[type_key] = type_obj

                    brand = brand_cache.get(marque)
                    if brand is None:
                        brand, _ = Brand.objects.get_or_create(magasin=magasin, nom=marque)
                        brand_cache[marque] = brand

                    prix_achat_val = prix_achat if prix_achat not in (None, "") else 0
                    prix_vente_val = prix_vente if prix_vente not in (None, "") else 0
                    actif_val = str(actif or "").strip().lower() not in ("non", "false", "0")

                    ref, ref_created = ProductReference.objects.get_or_create(
                        type=type_obj, brand=brand, reference_name=reference_name,
                        defaults={"prix_achat": prix_achat_val, "prix_vente": prix_vente_val, "actif": actif_val},
                    )
                    if ref_created:
                        created_refs += 1
                    else:
                        ref.prix_achat = prix_achat_val
                        ref.prix_vente = prix_vente_val
                        ref.actif = actif_val
                        ref.save(update_fields=["prix_achat", "prix_vente", "actif"])
                        updated_refs += 1

                    if not couleur:
                        continue

                    seuil_val = int(seuil_alerte) if seuil_alerte not in (None, "") else 1
                    stock_val = int(stock_actuel) if stock_actuel not in (None, "") else 0

                    variant, variant_created = ProductVariant.objects.get_or_create(
                        product_reference=ref, couleur=couleur,
                        defaults={"seuil_alerte": seuil_val},
                    )
                    if variant_created:
                        created_variants += 1
                        if stock_val > 0:
                            services.apply_stock_movement(
                                product_variant=variant, movement_type="ENTREE", quantite=stock_val,
                                origine="AJUSTEMENT", user=request.user, note="Import Excel",
                            )
                    else:
                        if variant.seuil_alerte != seuil_val:
                            variant.seuil_alerte = seuil_val
                            variant.save(update_fields=["seuil_alerte"])
                        diff = stock_val - variant.stock_actuel
                        if diff > 0:
                            services.apply_stock_movement(
                                product_variant=variant, movement_type="ENTREE", quantite=diff,
                                origine="AJUSTEMENT", user=request.user, note="Import Excel",
                            )
                        elif diff < 0:
                            services.apply_stock_movement(
                                product_variant=variant, movement_type="SORTIE", quantite=-diff,
                                origine="AJUSTEMENT", user=request.user, note="Import Excel",
                            )
                        updated_variants += 1
            except Exception as exc:
                errors.append(f"Ligne {row_index} : {exc}")

        return Response({
            "created_references": created_refs,
            "updated_references": updated_refs,
            "created_variants": created_variants,
            "updated_variants": updated_variants,
            "errors": errors,
        })


class ProductVariantViewSet(viewsets.ModelViewSet):
    serializer_class = ProductVariantSerializer
    permission_classes = [IsGerantOrReadOnly]
    http_method_names = ["get", "post", "delete", "head", "options"]  # stock modifiable via /adjust/, pas ici

    def get_queryset(self):
        qs = ProductVariant.objects.select_related("product_reference", "product_reference__brand").filter(
            product_reference__type__category__magasin__in=get_accessible_magasins(self.request.user)
        )
        reference_id = self.request.query_params.get("reference")
        if reference_id:
            qs = qs.filter(product_reference_id=reference_id)
        return qs

    def perform_create(self, serializer):
        # `stock_actuel` est en lecture seule sur le serializer (le stock ne
        # se modifie jamais hors mouvement tracé, §10 Smartreadme.md) — une
        # nouvelle couleur est donc créée à 0, puis un mouvement d'entrée est
        # appliqué si un stock initial a été demandé, pour garder la trace.
        variant = serializer.save()
        try:
            initial_stock = int(self.request.data.get("stock_actuel") or 0)
        except (TypeError, ValueError):
            initial_stock = 0
        if initial_stock > 0:
            services.apply_stock_movement(
                product_variant=variant,
                movement_type="ENTREE",
                quantite=initial_stock,
                origine="AJUSTEMENT",
                user=self.request.user,
                note="Stock initial à la création de la couleur",
            )
            variant.refresh_from_db()

    @action(detail=True, methods=["post"])
    def adjust(self, request, pk=None):
        """Ajustement manuel du stock — réservé au Gérant (§7.4 Smartreadme.md :
        'Modification manuelle du stock possible par le gérant uniquement')."""
        variant = self.get_object()
        serializer = StockAdjustmentSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        services.apply_stock_movement(
            product_variant=variant,
            movement_type=data["type"],
            quantite=data["quantite"],
            origine="AJUSTEMENT",
            user=request.user,
            note=data.get("note") or None,
        )
        variant.refresh_from_db()
        return Response(self.get_serializer(variant).data)


class StockMovementViewSet(viewsets.ReadOnlyModelViewSet):
    """Historique des mouvements de stock (§7.4/§10 Smartreadme.md) — lecture
    seule, toute écriture passe par catalog.services.apply_stock_movement."""

    serializer_class = StockMovementSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = StockMovement.objects.select_related(
            "product_variant", "product_variant__product_reference", "user"
        ).filter(
            product_variant__product_reference__type__category__magasin__in=get_accessible_magasins(
                self.request.user
            )
        )
        variant_id = self.request.query_params.get("variant")
        if variant_id:
            qs = qs.filter(product_variant_id=variant_id)
        return qs
