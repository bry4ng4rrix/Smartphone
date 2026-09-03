import io
import json
import os
import shutil
import tempfile
import time
import zipfile

import psutil
import openpyxl

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework import status, viewsets, serializers
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from django.db.models import Sum, F, DecimalField, Avg, Count, Value, Q
from django.db.models.functions import Coalesce
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.utils.text import slugify
from django.http import HttpResponse
from django.conf import settings
from django.core.management import call_command
from datetime import timedelta
from decimal import Decimal, InvalidOperation

from .models import CustomUser, MagasinProfile, EmployerProfile, AdminProfile, CaisseSession, CaisseMovement, ChatMessage, Notification, Subscription, LoginEvent, PlatformRequest, Device, SubscriptionOffer, EmployeePasswordResetRequest
from .serializers import RegisterSerializer, CaisseSessionSerializer, CaisseMovementSerializer, NotificationSerializer, MagasinProfileSerializer, ChatMessageSerializer, CompanySubscriptionSerializer, LoginEventSerializer, PlatformRequestSerializer, DeviceSerializer, SubscriptionOfferSerializer, EmployeePasswordResetRequestSerializer
from .permissions import IsAdmin, IsPlatformOwner, IsCompanyOwner, IsGerant, get_accessible_magasins, resolve_magasin_for_request, user_commande_role
from .subscriptions import get_company_magasins, get_company_user_ids, get_company_devices, get_subscription_owner, get_subscription, get_device_limit_info, parse_device_name
from rest_framework_simplejwt.views import TokenViewBase
from .authentication import CustomTokenObtainPairSerializer

from catalog.models import Brand, ProductCategory, ProductReference, ProductType, ProductVariant, StockMovement
from catalog.services import apply_stock_movement
from orders.models import Order, OrderItem, OrderStatusHistory
from suppliers.models import SupplierOrderLine


def is_company_owner(user):
    """True only for the admin who actually owns the company (has an
    AdminProfile). Co-admins added via AddAdminView share full data access to
    the company's magasins/products but are never owners of the company
    itself — company-level actions (adding/removing admins, subscription,
    devices) stay reserved to this one account."""
    return bool(
        user and user.is_authenticated and user.role == "admin"
        and AdminProfile.objects.filter(user=user).exists()
    )


def get_user_admin(user):
    if not user or user.is_anonymous:
        return None
    if user.role == "admin":
        return user
    elif user.role == "magasin":
        try:
            return user.magasin_profile.admin
        except Exception:
            return None
    elif user.role == "employer":
        try:
            ep = user.employer_profile
            if ep.admin:
                return ep.admin
            if ep.magasin:
                return ep.magasin.admin
        except Exception:
            return None
    return None


def get_company_magasins(user):
    """Magasins forming the 'company' a user belongs to.

    Unlike get_user_admin (which only resolves the single primary/owner admin),
    this also accounts for co-admins linked through MagasinProfile.admins (M2M),
    so admins added via AddAdminView are recognized as part of the same company.
    """
    if not user or user.is_anonymous:
        return MagasinProfile.objects.none()
    if user.role == "admin":
        return MagasinProfile.objects.filter(Q(admin=user) | Q(admins=user)).distinct()
    elif user.role == "magasin":
        try:
            return MagasinProfile.objects.filter(id=user.magasin_profile.id)
        except Exception:
            return MagasinProfile.objects.none()
    elif user.role == "employer":
        try:
            ep = user.employer_profile
            if ep.magasin:
                return MagasinProfile.objects.filter(id=ep.magasin.id)
        except Exception:
            pass
    return MagasinProfile.objects.none()


def get_company_id(user):
    """Stable identifier for a user's company, used to scope the general chat room.

    Based on the company's primary magasin owner (admin FK), so every co-admin,
    gérant and employer of the same magasins lands in the same room regardless
    of which admin is currently logged in.
    """
    magasin = get_company_magasins(user).order_by("id").first()
    return magasin.admin_id if magasin else None


def _stock_value_for_magasins(magasins):
    """Valeur du stock au prix de vente catalogue — le coût d'achat n'est
    plus suivi au niveau du catalogue depuis le passage au modèle
    Smartreadme.md (seul `prix_vente` y figure, §8.1)."""
    return ProductVariant.objects.filter(
        product_reference__type__category__magasin__in=magasins
    ).aggregate(
        total=Coalesce(
            Sum(F('stock_actuel') * F('product_reference__prix_vente'), output_field=DecimalField()),
            0,
            output_field=DecimalField(),
        )
    )['total']


def _cost_by_variant(magasins):
    """Coût moyen connu par variante — moyenne des coûts unitaires des
    lignes de commande fournisseur reçues (§7.6 Smartreadme.md). Meilleure
    estimation disponible du coût de revient, faute de suivi de coût par
    vente individuelle (le catalogue ne porte que le prix de vente)."""
    rows = SupplierOrderLine.objects.filter(
        supplier_order__statut="RECU", supplier_order__magasin__in=magasins
    ).values("product_variant_id").annotate(avg_cost=Avg("cout_unitaire_calcule"))
    return {row["product_variant_id"]: row["avg_cost"] or 0 for row in rows}


def _order_items_revenue_and_profit(orders_qs, cost_map):
    """CA produits (hors frais livraison) et bénéfice estimé pour un
    queryset de commandes (normalement filtré sur statut_courant='LIVRE'),
    à partir du coût moyen connu par variante (0 si jamais reçu via une
    commande fournisseur — le bénéfice est alors surestimé pour ces lignes)."""
    total_revenue = 0
    total_cost = 0
    for item in OrderItem.objects.filter(order__in=orders_qs).select_related("product_variant"):
        total_revenue += item.prix_unitaire * item.quantite
        total_cost += cost_map.get(item.product_variant_id, 0) * item.quantite
    return total_revenue, total_revenue - total_cost


def _top_bottom_products(magasin_ids, livrees_qs, limit=5):
    """(top, bottom) N références par quantité vendue (commandes Livrées) —
    agrégation faite en Python (§7.7 Smartreadme.md : TOP produits) : le
    volume reste modeste à l'échelle d'un magasin, et ça évite une jointure
    multi-niveaux fragile pour inclure les références jamais vendues dans
    le classement 'bottom'."""
    items = OrderItem.objects.filter(order__in=livrees_qs).select_related("product_variant__product_reference")
    qty_by_ref_id = {}
    ref_names = {}
    for item in items:
        ref = item.product_variant.product_reference
        qty_by_ref_id[ref.id] = qty_by_ref_id.get(ref.id, 0) + item.quantite
        ref_names[ref.id] = ref.reference_name

    top = sorted(
        ({"name": ref_names[rid], "qty_sold": qty} for rid, qty in qty_by_ref_id.items()),
        key=lambda r: -r["qty_sold"],
    )[:limit]

    all_refs = ProductReference.objects.filter(type__category__magasin_id__in=magasin_ids).values_list("id", "reference_name")
    bottom = sorted(
        ({"name": name, "qty_sold": qty_by_ref_id.get(rid, 0)} for rid, name in all_refs),
        key=lambda r: r["qty_sold"],
    )[:limit]
    return top, bottom


class AdminMagasinOverviewView(APIView):
    """Résumé par magasin pour l'admin (valeur de stock, bénéfice estimé,
    nb produits/ventes/employés) — alimente les cartes 'Meilleurs magasins'
    du dashboard (dasboard.md)."""

    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        week_start = timezone.now() - timedelta(days=7)
        magasins = MagasinProfile.objects.filter(Q(admin=request.user) | Q(admins=request.user)).distinct()
        response_data = []

        for magasin in magasins:
            magasin_qs = MagasinProfile.objects.filter(id=magasin.id)
            livrees_qs = Order.objects.filter(magasin=magasin, statut_courant="LIVRE")

            cost_map = _cost_by_variant(magasin_qs)
            _, total_profit = _order_items_revenue_and_profit(livrees_qs, cost_map)

            response_data.append({
                "magasin_id": magasin.id,
                "shop_name": magasin.shop_name,
                "total_stock_value": _stock_value_for_magasins(magasin_qs),
                "total_profit": total_profit,
                "number_of_products": ProductReference.objects.filter(type__category__magasin=magasin).count(),
                "number_of_sales_week": livrees_qs.filter(updated_at__gte=week_start).count(),
                "number_of_employees": EmployerProfile.objects.filter(magasin=magasin).count(),
            })

        return Response({"magasins": response_data})


# =========================
# AUTH LOGIN
# =========================
class CustomLoginView(TokenViewBase):
    serializer_class = CustomTokenObtainPairSerializer

# =========================
# REGISTER
# =========================
class RegisterView(APIView):
    def post(self, request):
        serializer = RegisterSerializer(data=request.data, context={"request": request})
        if serializer.is_valid():
            user = serializer.save()
            # If the new user is an admin, create a default "Stock Local" store linked to this admin.
            if user.role == "admin":
                from .models import MagasinProfile
                magasin = MagasinProfile.objects.create(
                    admin=user,
                    shop_name="Stock Local",
                    description="Magasin pour les stocks locaux",
                )
                magasin.admins.add(user)
            return Response({"message": "Inscription réussie", "id": user.id})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class AddAdminView(APIView):
    # Only the company's owner (the admin with an AdminProfile) may add a
    # co-admin — a co-admin adding further co-admins would let access spread
    # without the founder's knowledge.
    permission_classes = [IsAuthenticated, IsCompanyOwner]

    def post(self, request):
        """Create a new admin user and associate with existing magasins of the requester."""
        serializer = RegisterSerializer(data=request.data)
        if serializer.is_valid():
            new_admin = serializer.save()
            # Ensure role is admin
            new_admin.role = "admin"
            new_admin.save()
            # This is a co-admin on an EXISTING company, never a new one:
            # RegisterSerializer always creates an AdminProfile+Subscription for
            # role="admin", but that would spuriously appear as its own company
            # in the platform dashboard. Drop it (cascades to the Subscription).
            AdminProfile.objects.filter(user=new_admin).delete()
            # Give the new admin the exact same access as the creator: every magasin
            # where the requester is the owner (admin FK) or a co-admin (admins M2M).
            current_admin = request.user
            magasins = MagasinProfile.objects.filter(Q(admin=current_admin) | Q(admins=current_admin)).distinct()
            for magasin in magasins:
                magasin.admins.add(new_admin)
            return Response({"message": "Admin ajouté", "id": new_admin.id}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# =========================
# APPROVE USER
# =========================
class ApproveUserView(APIView):
    permission_classes = [IsAuthenticated]
    def put(self, request, user_id):
        current_user = request.user
        if current_user.role not in ["admin", "magasin"]:
            return Response({"error": "Permission refusée"}, status=403)

        user_admin = get_user_admin(current_user)
        if not user_admin:
            return Response({"error": "Permission refusée : entreprise introuvable."}, status=403)

        if current_user.role == "magasin":
            try:
                magasin = current_user.magasin_profile
                is_member = CustomUser.objects.filter(
                    id=user_id,
                    role="employer",
                    employer_profile__magasin=magasin
                ).exists()
                if not is_member:
                    return Response({"error": "Permission refusée : cet employé n'appartient pas à votre magasin."}, status=403)
            except Exception:
                return Response({"error": "Magasin introuvable"}, status=404)
        else: # admin
            is_member = CustomUser.objects.filter(
                Q(id=user_id),
                Q(magasin_profile__admin=user_admin) | Q(employer_profile__admin=user_admin) | Q(employer_profile__magasin__admin=user_admin)
            ).exists()
            if not is_member:
                return Response({"error": "Permission refusée : cet utilisateur n'appartient pas à votre entreprise."}, status=403)

        try:
            user = CustomUser.objects.get(id=user_id)
            user.is_confirmed = True
            user.save()
            return Response({"message": "Utilisateur approuvé"})
        except CustomUser.DoesNotExist:
            return Response({"error": "Utilisateur introuvable"}, status=404)

# =========================
# MY COMPANY (tenant side: devices, subscription, requests to Label Technology)
# =========================
class MyCompanyDevicesView(APIView):
    # Owner-only: device/security management is a company-ownership concern,
    # not something a co-admin should manage on the founder's behalf.
    permission_classes = [IsAuthenticated, IsCompanyOwner]

    def get(self, request):
        owner = get_subscription_owner(request.user)
        if not owner:
            return Response({"error": "Société introuvable"}, status=404)
        devices = Device.objects.filter(admin_profile=owner.admin_profile).select_related("user")
        count, limit = get_device_limit_info(request.user)
        return Response({
            "devices": DeviceSerializer(devices, many=True).data,
            "count": count,
            "limit": limit,
        })


class MyCompanySubscriptionView(APIView):
    # Owner-only: billing/subscription is a company-ownership concern.
    permission_classes = [IsAuthenticated, IsCompanyOwner]

    def get(self, request):
        owner = get_subscription_owner(request.user)
        sub = get_subscription(request.user)
        if not sub:
            return Response({"status": "pending", "trial_ends_at": None, "is_currently_active": False, "days_left_in_trial": None})
        return Response({
            "status": sub.status,
            "trial_ends_at": sub.trial_ends_at,
            "is_currently_active": sub.is_currently_active,
            "days_left_in_trial": sub.days_left_in_trial,
            "offer": SubscriptionOfferSerializer(sub.offer).data if sub.offer else None,
        })


class MyCompanyRequestsView(APIView):
    # Owner-only: requests to Label Technology (activation, device removal)
    # are a company-ownership concern.
    permission_classes = [IsAuthenticated, IsCompanyOwner]

    def get(self, request):
        owner = get_subscription_owner(request.user)
        if not owner:
            return Response({"error": "Société introuvable"}, status=404)
        reqs = PlatformRequest.objects.filter(admin_profile=owner.admin_profile).order_by("-created_at")
        return Response(PlatformRequestSerializer(reqs, many=True).data)

    def post(self, request):
        owner = get_subscription_owner(request.user)
        if not owner:
            return Response({"error": "Société introuvable"}, status=404)

        request_type = request.data.get("request_type")
        if request_type not in dict(PlatformRequest.REQUEST_TYPES):
            return Response({"error": "Type de demande invalide."}, status=400)

        login_event = None
        device = None
        if request_type == "device_deletion":
            device_id = request.data.get("device_id")
            login_event_id = request.data.get("login_event_id")
            if device_id:
                device = Device.objects.filter(id=device_id, admin_profile=owner.admin_profile).first()
                if not device:
                    return Response({"error": "Appareil introuvable."}, status=404)
                if PlatformRequest.objects.filter(device=device, status="pending").exists():
                    return Response({"error": "Une demande est déjà en attente pour cet appareil."}, status=400)
            else:
                login_event = LoginEvent.objects.filter(id=login_event_id).first()
                if not login_event or login_event.user_id not in get_company_user_ids(owner):
                    return Response({"error": "Appareil introuvable."}, status=404)
                if PlatformRequest.objects.filter(login_event=login_event, status="pending").exists():
                    return Response({"error": "Une demande est déjà en attente pour cet appareil."}, status=400)

        if request_type == "activation":
            if PlatformRequest.objects.filter(
                admin_profile=owner.admin_profile, request_type="activation", status="pending"
            ).exists():
                return Response({"error": "Une demande d'activation est déjà en attente."}, status=400)

        req = PlatformRequest.objects.create(
            request_type=request_type,
            admin_profile=owner.admin_profile,
            requested_by=request.user,
            login_event=login_event,
            device=device,
            note=request.data.get("note", ""),
        )
        return Response(PlatformRequestSerializer(req).data, status=201)


# =========================
# PLATFORM ADMIN (Label Technology)
# =========================
class PlatformCompanyListView(APIView):
    permission_classes = [IsAuthenticated, IsPlatformOwner]

    def get(self, request):
        qs = AdminProfile.objects.select_related("user", "subscription").order_by("company_name")
        serializer = CompanySubscriptionSerializer(qs, many=True, context={"request": request})
        return Response(serializer.data)

    def post(self, request):
        data = request.data.copy()
        data["role"] = "admin"
        data.setdefault("username", data.get("email"))
        serializer = RegisterSerializer(data=data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)

        user = serializer.save()
        user.role = "admin"
        user.save()

        magasin = MagasinProfile.objects.create(
            admin=user, shop_name="Stock Local", description="Magasin pour les stocks locaux"
        )
        magasin.admins.add(user)

        initial_status = request.data.get("status", "pending")
        if initial_status not in dict(Subscription.STATUS_CHOICES):
            initial_status = "pending"

        sub = user.admin_profile.subscription
        sub.status = initial_status
        sub.trial_ends_at = timezone.now() + timedelta(days=30) if initial_status == "trial" else None
        sub.updated_by = request.user
        sub.save()

        out = CompanySubscriptionSerializer(user.admin_profile, context={"request": request})
        return Response(out.data, status=201)


class PlatformCompanyStatusUpdateView(APIView):
    permission_classes = [IsAuthenticated, IsPlatformOwner]

    def patch(self, request, admin_profile_id):
        try:
            admin_profile = AdminProfile.objects.get(id=admin_profile_id)
        except AdminProfile.DoesNotExist:
            return Response({"error": "Société introuvable"}, status=404)

        new_status = request.data.get("status")
        if new_status not in dict(Subscription.STATUS_CHOICES):
            return Response({"status": "Statut invalide."}, status=400)

        sub, _ = Subscription.objects.get_or_create(admin_profile=admin_profile)
        sub.status = new_status
        sub.trial_ends_at = timezone.now() + timedelta(days=30) if new_status == "trial" else None
        sub.updated_by = request.user
        sub.save()

        serializer = CompanySubscriptionSerializer(admin_profile, context={"request": request})
        return Response(serializer.data)


class PlatformActivateAllView(APIView):
    permission_classes = [IsAuthenticated, IsPlatformOwner]

    def post(self, request):
        missing = AdminProfile.objects.filter(subscription__isnull=True)
        Subscription.objects.bulk_create([
            Subscription(admin_profile=ap, status="active", updated_by=request.user)
            for ap in missing
        ])
        count = Subscription.objects.exclude(status="active").update(
            status="active", trial_ends_at=None, updated_by=request.user, updated_at=timezone.now()
        )
        return Response({"activated": count})


class PlatformCompanyDetailView(APIView):
    permission_classes = [IsAuthenticated, IsPlatformOwner]

    def patch(self, request, admin_profile_id):
        try:
            admin_profile = AdminProfile.objects.get(id=admin_profile_id)
        except AdminProfile.DoesNotExist:
            return Response({"error": "Société introuvable"}, status=404)

        company_name = request.data.get("company_name")
        full_name = request.data.get("admin_full_name")
        phone = request.data.get("admin_phone")

        if company_name:
            admin_profile.company_name = company_name
            admin_profile.save()

        if full_name or phone is not None:
            user = admin_profile.user
            if full_name:
                user.full_name = full_name
            if phone is not None:
                user.phone = phone
            user.save()

        out = CompanySubscriptionSerializer(admin_profile, context={"request": request})
        return Response(out.data)

    def delete(self, request, admin_profile_id):
        try:
            admin_profile = AdminProfile.objects.get(id=admin_profile_id)
        except AdminProfile.DoesNotExist:
            return Response({"error": "Société introuvable"}, status=404)

        buffer, filename = _build_company_backup_zip(admin_profile)

        admin_user = admin_profile.user
        magasins = list(get_company_magasins(admin_user))
        user_ids = get_company_user_ids(admin_user)

        for mp in magasins:
            mp.delete()
        CustomUser.objects.filter(id__in=user_ids).delete()

        response = HttpResponse(buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


def _build_company_backup_zip(admin_profile):
    """Build a zip (data.json + donnees.xlsx + media/) containing every record
    belonging to this company. Returns (BytesIO, filename)."""
    from catalog.serializers import ProductVariantSerializer, StockMovementSerializer
    from orders.serializers import OrderGerantSerializer

    admin_user = admin_profile.user
    magasins = get_company_magasins(admin_user)
    user_ids = get_company_user_ids(admin_user)

    variants = ProductVariant.objects.filter(product_reference__type__category__magasin__in=magasins)
    orders_qs = Order.objects.filter(magasin__in=magasins)
    movements = StockMovement.objects.filter(product_variant__product_reference__type__category__magasin__in=magasins)
    users = CustomUser.objects.filter(id__in=user_ids)

    payload = {
        "company_name": admin_profile.company_name,
        "exported_at": timezone.now().isoformat(),
        "products": ProductVariantSerializer(variants, many=True).data,
        "orders": OrderGerantSerializer(orders_qs, many=True).data,
        "movements": StockMovementSerializer(movements, many=True).data,
        "magasins": MagasinProfileSerializer(magasins, many=True).data,
        "users": [
            {
                "id": u.id, "full_name": u.full_name, "email": u.email,
                "role": u.role, "phone": u.phone, "is_confirmed": u.is_confirmed,
            }
            for u in users
        ],
    }

    wb = openpyxl.Workbook()
    ws_products = wb.active
    ws_products.title = "Produits"
    ws_products.append(["ID", "Référence", "Marque", "Couleur", "Stock", "Seuil alerte", "Prix vente"])
    for v in variants.select_related("product_reference", "product_reference__brand"):
        ws_products.append([
            v.id, v.product_reference.reference_name, v.product_reference.brand.nom, v.couleur,
            v.stock_actuel, v.seuil_alerte, float(v.product_reference.prix_vente or 0),
        ])

    ws_orders = wb.create_sheet("Commandes")
    ws_orders.append(["N°", "Client", "Téléphone", "Zone", "Statut", "Total à payer", "Magasin", "Date"])
    for o in orders_qs.select_related("magasin"):
        ws_orders.append([
            o.numero, o.client_nom, o.telephone, o.get_livraison_zone_display(), o.get_statut_courant_display(),
            float(o.total_a_payer or 0), o.magasin.shop_name if o.magasin else "",
            o.date_commande.isoformat() if o.date_commande else "",
        ])

    ws_movements = wb.create_sheet("Mouvements")
    ws_movements.append(["ID", "Produit", "Type", "Quantité", "Origine", "Utilisateur", "Date"])
    for m in movements.select_related("product_variant", "product_variant__product_reference", "user"):
        ws_movements.append([
            m.id, str(m.product_variant), m.get_type_display(), m.quantite, m.get_origine_display(),
            m.user.full_name if m.user else "",
            m.timestamp.isoformat() if m.timestamp else "",
        ])

    ws_users = wb.create_sheet("Utilisateurs")
    ws_users.append(["ID", "Nom", "Email", "Rôle", "Téléphone", "Confirmé"])
    for u in users:
        ws_users.append([u.id, u.full_name, u.email, u.role, u.phone or "", u.is_confirmed])

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("data.json", json.dumps(payload, indent=2, default=str))
        zf.writestr("donnees.xlsx", excel_buffer.getvalue())

        image_fields = []
        if admin_profile.logo:
            image_fields.append(admin_profile.logo)
        for mp in magasins:
            if mp.shop_logo:
                image_fields.append(mp.shop_logo)

        for field in image_fields:
            try:
                if os.path.isfile(field.path):
                    zf.write(field.path, os.path.join("media", field.name))
            except Exception:
                continue

    buffer.seek(0)
    slug = slugify(admin_profile.company_name) or f"societe-{admin_profile.id}"
    filename = f"{slug}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.zip"
    return buffer, filename


class PlatformCompanyBackupView(APIView):
    permission_classes = [IsAuthenticated, IsPlatformOwner]

    def get(self, request, admin_profile_id):
        try:
            admin_profile = AdminProfile.objects.get(id=admin_profile_id)
        except AdminProfile.DoesNotExist:
            return Response({"error": "Société introuvable"}, status=404)

        buffer, filename = _build_company_backup_zip(admin_profile)
        response = HttpResponse(buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class PlatformCompanyDevicesView(APIView):
    permission_classes = [IsAuthenticated, IsPlatformOwner]

    def get(self, request, admin_profile_id):
        try:
            admin_profile = AdminProfile.objects.get(id=admin_profile_id)
        except AdminProfile.DoesNotExist:
            return Response({"error": "Société introuvable"}, status=404)

        devices = Device.objects.filter(admin_profile=admin_profile).select_related("user")
        count, limit = get_device_limit_info(admin_profile.user)
        return Response({
            "devices": DeviceSerializer(devices, many=True).data,
            "count": count,
            "limit": limit,
        })

    def delete(self, request, admin_profile_id):
        device_id = request.query_params.get("device_id") or request.data.get("device_id")
        try:
            admin_profile = AdminProfile.objects.get(id=admin_profile_id)
        except AdminProfile.DoesNotExist:
            return Response({"error": "Société introuvable"}, status=404)

        device = Device.objects.filter(id=device_id, admin_profile=admin_profile).first()
        if not device:
            return Response({"error": "Appareil introuvable."}, status=404)
        device.delete()
        return Response(status=204)


class PlatformOfferListView(APIView):
    permission_classes = [IsAuthenticated, IsPlatformOwner]

    def get(self, request):
        offers = SubscriptionOffer.objects.all()
        return Response(SubscriptionOfferSerializer(offers, many=True).data)

    def post(self, request):
        serializer = SubscriptionOfferSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)
        serializer.save()
        return Response(serializer.data, status=201)


class PlatformOfferDetailView(APIView):
    permission_classes = [IsAuthenticated, IsPlatformOwner]

    def patch(self, request, offer_id):
        try:
            offer = SubscriptionOffer.objects.get(id=offer_id)
        except SubscriptionOffer.DoesNotExist:
            return Response({"error": "Offre introuvable"}, status=404)

        serializer = SubscriptionOfferSerializer(offer, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)
        serializer.save()
        return Response(serializer.data)

    def delete(self, request, offer_id):
        try:
            offer = SubscriptionOffer.objects.get(id=offer_id)
        except SubscriptionOffer.DoesNotExist:
            return Response({"error": "Offre introuvable"}, status=404)
        offer.delete()
        return Response(status=204)


class PlatformCompanyOfferAssignView(APIView):
    permission_classes = [IsAuthenticated, IsPlatformOwner]

    def patch(self, request, admin_profile_id):
        try:
            admin_profile = AdminProfile.objects.get(id=admin_profile_id)
        except AdminProfile.DoesNotExist:
            return Response({"error": "Société introuvable"}, status=404)

        offer_id = request.data.get("offer_id")
        offer = None
        if offer_id:
            offer = SubscriptionOffer.objects.filter(id=offer_id).first()
            if not offer:
                return Response({"error": "Offre introuvable"}, status=404)

        sub, _ = Subscription.objects.get_or_create(admin_profile=admin_profile)
        sub.offer = offer
        sub.updated_by = request.user
        sub.save()

        return Response(CompanySubscriptionSerializer(admin_profile, context={"request": request}).data)


class PlatformMonitoringView(APIView):
    permission_classes = [IsAuthenticated, IsPlatformOwner]

    def get(self, request):
        t0 = time.monotonic()
        cpu = psutil.cpu_percent(interval=0.1)
        mem = psutil.virtual_memory()
        disk = shutil.disk_usage(settings.BASE_DIR)

        db_engine = settings.DATABASES["default"]["ENGINE"]
        db_size = None
        if db_engine.endswith("sqlite3"):
            db_path = settings.DATABASES["default"]["NAME"]
            if os.path.isfile(db_path):
                db_size = os.path.getsize(db_path)

        return Response({
            "cpu_percent": cpu,
            "ram_used": mem.used,
            "ram_total": mem.total,
            "ram_percent": mem.percent,
            "disk_used": disk.used,
            "disk_total": disk.total,
            "disk_percent": round(disk.used / disk.total * 100, 1) if disk.total else 0,
            "db_size_bytes": db_size,
            "server_time": timezone.now().isoformat(),
            "process_time_ms": round((time.monotonic() - t0) * 1000, 2),
        })


def _activate_subscription(admin_profile, updated_by, offer=None):
    """Activates (or reactivates) a company's subscription, optionally
    assigning a SubscriptionOffer plan. Shared by manual approval
    (PlatformRequestResolveView) and the simulated payment endpoint
    (PublicPaymentRequestView)."""
    sub, _ = Subscription.objects.get_or_create(admin_profile=admin_profile)
    sub.status = "active"
    sub.trial_ends_at = None
    if offer is not None:
        sub.offer = offer
    sub.updated_by = updated_by
    sub.save()
    return sub


class PlatformRequestListView(APIView):
    permission_classes = [IsAuthenticated, IsPlatformOwner]

    def get(self, request):
        qs = PlatformRequest.objects.select_related("admin_profile", "requested_by", "login_event")
        status_filter = request.query_params.get("status")
        if status_filter:
            qs = qs.filter(status=status_filter)
        return Response(PlatformRequestSerializer(qs, many=True).data)


class PlatformRequestResolveView(APIView):
    permission_classes = [IsAuthenticated, IsPlatformOwner]

    def patch(self, request, request_id):
        try:
            req = PlatformRequest.objects.get(id=request_id)
        except PlatformRequest.DoesNotExist:
            return Response({"error": "Demande introuvable"}, status=404)

        action = request.data.get("action")
        if action not in ("approve", "reject"):
            return Response({"error": "Action invalide."}, status=400)
        if req.status != "pending":
            return Response({"error": "Cette demande a déjà été traitée."}, status=400)

        if action == "approve":
            if req.request_type == "device_deletion" and req.device:
                req.device.delete()
                req.device = None
            elif req.request_type == "device_deletion" and req.login_event:
                req.login_event.delete()
                req.login_event = None
            elif req.request_type == "activation":
                _activate_subscription(req.admin_profile, request.user)
            elif req.request_type == "payment":
                _activate_subscription(req.admin_profile, request.user, offer=req.offer)
            req.status = "approved"
        else:
            req.status = "rejected"

        req.resolved_by = request.user
        req.resolved_at = timezone.now()
        req.save()
        return Response(PlatformRequestSerializer(req).data)


# =========================
# EMPLOYEE PASSWORD RESET (magasin/employer -> admin)
# =========================
class EmployeePasswordResetListView(APIView):
    """Lists forgot-password requests from the admin's magasin/employer
    accounts, for the admin to approve or reject."""
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        qs = EmployeePasswordResetRequest.objects.filter(admin=request.user).select_related("user", "magasin")
        status_filter = request.query_params.get("status")
        if status_filter:
            qs = qs.filter(status=status_filter)
        return Response(EmployeePasswordResetRequestSerializer(qs, many=True).data)


class EmployeePasswordResetResolveView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]

    def patch(self, request, request_id):
        try:
            req = EmployeePasswordResetRequest.objects.get(id=request_id, admin=request.user)
        except EmployeePasswordResetRequest.DoesNotExist:
            return Response({"error": "Demande introuvable"}, status=404)

        action = request.data.get("action")
        if action not in ("approve", "reject"):
            return Response({"error": "Action invalide."}, status=400)
        if req.status != "pending":
            return Response({"error": "Cette demande a déjà été traitée."}, status=400)

        req.status = "approved" if action == "approve" else "rejected"
        req.resolved_by = request.user
        req.resolved_at = timezone.now()
        req.save()
        return Response(EmployeePasswordResetRequestSerializer(req).data)


class PlatformExpiringSoonView(APIView):
    permission_classes = [IsAuthenticated, IsPlatformOwner]
    THRESHOLD_DAYS = 3

    def get(self, request):
        qs = AdminProfile.objects.select_related("user", "subscription").filter(
            subscription__status="trial"
        )
        expiring = [
            ap for ap in qs
            if ap.subscription.days_left_in_trial is not None
            and ap.subscription.days_left_in_trial <= self.THRESHOLD_DAYS
        ]
        out = CompanySubscriptionSerializer(expiring, many=True, context={"request": request})
        return Response(out.data)


# =========================
# PUBLIC (unauthenticated — used by the /abonnement-expire page, since a
# blocked/expired subscription means the visitor has no valid JWT)
# =========================
class PublicOfferListView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        offers = SubscriptionOffer.objects.filter(is_active=True)
        return Response(SubscriptionOfferSerializer(offers, many=True).data)


def _check_credentials(email, password):
    """Verifies email/password without going through the real login endpoint
    (which would itself reject a blocked/expired subscription). Returns
    (user, error_response) — error_response is a ready-to-return Response
    when the credentials or account are invalid, otherwise None.
    """
    email = (email or "").strip().lower()
    user = CustomUser.objects.filter(email__iexact=email).first()
    if not user or not user.check_password(password or ""):
        return None, Response({"error": "Email ou mot de passe incorrect."}, status=401)
    if not user.is_confirmed:
        return None, Response({"error": "Compte non approuvé. Contactez votre administrateur."}, status=403)
    return user, None


class PublicVerifyAccountView(APIView):
    """Identifies who is trying to pay (company + user name) without
    issuing any JWT — the account's subscription is by definition inactive
    at this point, so no real session should be created here. Used by the
    /abonnement-expire payment flow to decide whether to show the payment
    form (admins only) or a "wait for your admin" message."""
    permission_classes = [AllowAny]

    def post(self, request):
        user, error = _check_credentials(request.data.get("email"), request.data.get("password"))
        if error:
            return error

        owner = get_subscription_owner(user)
        admin_profile = getattr(owner, "admin_profile", None) if owner else None
        if not admin_profile:
            return Response({"error": "Aucune société associée à ce compte."}, status=404)

        return Response({
            "email": user.email,
            "full_name": user.full_name,
            "role": user.role,
            "is_admin": user.role == "admin",
            "company_name": admin_profile.company_name,
        })


class PublicPaymentRequestView(APIView):
    """Records a direct-payment request from a blocked/expired company and
    (until a real payment gateway is wired in) immediately simulates its
    approval: activates the subscription with the chosen offer, exactly as
    a Label Technology admin approving the request manually would.

    Only the company's admin can trigger this (re-checked here server-side,
    not just gated in the UI) — credentials are re-verified independently of
    PublicVerifyAccountView, since that call only gates what the frontend
    displays and must not be trusted as proof of authorization on its own.

    TODO: once a real payment provider (Mvola/PayPal/Visa/Mastercard) is
    integrated, this must only create the pending PlatformRequest and defer
    activation to a verified payment webhook / manual review, instead of
    auto-approving on submit.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        user, error = _check_credentials(request.data.get("email"), request.data.get("password"))
        if error:
            return error

        if user.role != "admin":
            return Response(
                {"error": "Seul un administrateur de la société peut effectuer ce paiement."},
                status=403,
            )

        offer_id = request.data.get("offer_id")
        payment_method = request.data.get("payment_method")
        payment_reference = (request.data.get("payment_reference") or "").strip()
        if payment_method not in dict(PlatformRequest.PAYMENT_METHOD_CHOICES):
            return Response({"error": "Moyen de paiement invalide."}, status=400)

        if payment_method == "mvola":
            digits = payment_reference.replace(" ", "")
            if not digits.isdigit() or len(digits) < 9:
                return Response({"error": "Numéro de téléphone MVola invalide."}, status=400)
        elif payment_method == "paypal":
            if "@" not in payment_reference:
                return Response({"error": "Email PayPal invalide."}, status=400)
        else:  # visa / mastercard
            if len(payment_reference) < 2:
                return Response({"error": "Le nom du titulaire de la carte est obligatoire."}, status=400)

        offer = SubscriptionOffer.objects.filter(id=offer_id, is_active=True).first()
        if not offer:
            return Response({"error": "Offre introuvable."}, status=404)

        owner = get_subscription_owner(user)
        admin_profile = getattr(owner, "admin_profile", None) if owner else None
        if not admin_profile:
            return Response({"error": "Aucune société associée à ce compte."}, status=404)

        if PlatformRequest.objects.filter(admin_profile=admin_profile, request_type="payment", status="pending").exists():
            return Response({"error": "Une demande de paiement est déjà en attente pour cette société."}, status=400)

        req = PlatformRequest.objects.create(
            request_type="payment",
            admin_profile=admin_profile,
            requested_by=user,
            offer=offer,
            payment_method=payment_method,
            payment_reference=payment_reference,
            contact_email=user.email,
            note=request.data.get("note", ""),
        )

        # Simulated auto-approval: no real payment gateway yet (see docstring above).
        _activate_subscription(admin_profile, updated_by=None, offer=offer)
        req.status = "approved"
        req.resolved_at = timezone.now()
        req.save()

        return Response(PlatformRequestSerializer(req).data, status=201)


class PublicForgotPasswordRequestView(APIView):
    """Step 1 of the forgot-password flow (no auth, since the requester can't
    log in). Identifies the account by email and routes the request to the
    right approver: Label Technology for admin (société) accounts, or the
    société's admin for magasin/employer accounts. There is no email backend
    configured in this project, so no reset link is sent — the requester
    comes back later (see PublicForgotPasswordStatusView) to check whether
    the request was approved and set a new password themselves."""
    permission_classes = [AllowAny]

    def post(self, request):
        email = (request.data.get("email") or "").strip().lower()
        if not email:
            return Response({"error": "Email requis."}, status=400)

        user = CustomUser.objects.filter(email__iexact=email).first()
        if not user:
            return Response({"error": "Aucun compte avec cet email."}, status=404)

        if user.role == "admin":
            admin_profile = AdminProfile.objects.filter(user=user).first()
            if not admin_profile:
                return Response({"error": "Aucune société associée à ce compte."}, status=404)
            if PlatformRequest.objects.filter(
                admin_profile=admin_profile, request_type="password_reset", status="pending"
            ).exists():
                return Response({"error": "Une demande est déjà en attente."}, status=400)

            PlatformRequest.objects.create(
                request_type="password_reset",
                admin_profile=admin_profile,
                requested_by=user,
                contact_email=user.email,
            )
            return Response({
                "queue": "label",
                "message": "Votre demande a été transmise à Label Technology pour validation.",
            }, status=201)

        elif user.role in ("magasin", "employer"):
            admin = None
            magasin = None
            if user.role == "magasin":
                magasin = MagasinProfile.objects.filter(user=user).first()
                admin = magasin.admin if magasin else None
            else:
                employer_profile = EmployerProfile.objects.filter(user=user).first()
                if employer_profile:
                    magasin = employer_profile.magasin
                    admin = employer_profile.admin or (magasin.admin if magasin else None)

            if not admin:
                return Response({"error": "Aucun administrateur associé à ce compte."}, status=404)

            if EmployeePasswordResetRequest.objects.filter(user=user, status="pending").exists():
                return Response({"error": "Une demande est déjà en attente."}, status=400)

            EmployeePasswordResetRequest.objects.create(user=user, admin=admin, magasin=magasin)
            return Response({
                "queue": "admin",
                "message": "Votre demande a été transmise à votre administrateur pour validation.",
            }, status=201)

        return Response({"error": "Réinitialisation non disponible pour ce type de compte."}, status=400)


class PublicForgotPasswordStatusView(APIView):
    """Step 2: the requester comes back with their email to check whether
    their pending request has been approved yet."""
    permission_classes = [AllowAny]

    def get(self, request):
        email = (request.query_params.get("email") or "").strip().lower()
        user = CustomUser.objects.filter(email__iexact=email).first()
        if not user:
            return Response({"status": "none"})

        if user.role == "admin":
            req = PlatformRequest.objects.filter(
                requested_by=user, request_type="password_reset", consumed_at__isnull=True
            ).order_by("-created_at").first()
        else:
            req = EmployeePasswordResetRequest.objects.filter(
                user=user, consumed_at__isnull=True
            ).order_by("-created_at").first()

        if not req:
            return Response({"status": "none"})
        return Response({"status": req.status})


class PublicForgotPasswordConfirmView(APIView):
    """Step 3: once approved, the requester sets their new password directly
    (no auth — that's the whole point of forgot-password). The approved
    request is marked consumed so it can't be replayed."""
    permission_classes = [AllowAny]

    def post(self, request):
        email = (request.data.get("email") or "").strip().lower()
        new_password = request.data.get("new_password") or ""
        if len(new_password) < 6:
            return Response({"error": "Le mot de passe doit contenir au moins 6 caractères."}, status=400)

        user = CustomUser.objects.filter(email__iexact=email).first()
        if not user:
            return Response({"error": "Aucun compte avec cet email."}, status=404)

        if user.role == "admin":
            req = PlatformRequest.objects.filter(
                requested_by=user, request_type="password_reset", status="approved", consumed_at__isnull=True
            ).order_by("-created_at").first()
        else:
            req = EmployeePasswordResetRequest.objects.filter(
                user=user, status="approved", consumed_at__isnull=True
            ).order_by("-created_at").first()

        if not req:
            return Response({"error": "Aucune demande approuvée trouvée pour cet email."}, status=400)

        user.set_password(new_password)
        user.save()
        req.consumed_at = timezone.now()
        req.save()
        return Response({"message": "Mot de passe mis à jour avec succès."})


# =========================
# MY PROFILE
# =========================
class Myprofile(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        data = {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "full_name": user.full_name,
            "phone": user.phone,
            "role": user.role,
            # Rôle au sens du module Commande (§4 Smartreadme.md) : "GERANT"
            # (admin/magasin), "PREPARATEUR"/"LIVREUR" (employer avec ce
            # commande_role) ou null — consommé par le client Flutter mobile
            # (smartcross), qui ne connaît que ces 3 rôles.
            "role_commande": user_commande_role(user),
            "is_confirmed": user.is_confirmed,
            "is_company_owner": False,
        }
        if user.role == "admin":
            try:
                p = user.admin_profile
                data["company_name"] = p.company_name
                data["logo"] = request.build_absolute_uri(p.logo.url) if p.logo else None
                data["is_company_owner"] = True
            except AdminProfile.DoesNotExist:
                # Co-admin (added via AddAdminView): no AdminProfile of their
                # own — show the owner's société name/logo instead of blank.
                owner = get_subscription_owner(user)
                owner_profile = getattr(owner, "admin_profile", None) if owner else None
                if owner_profile:
                    data["company_name"] = owner_profile.company_name
                    data["logo"] = request.build_absolute_uri(owner_profile.logo.url) if owner_profile.logo else None
            # Commodité pour les clients mobiles (une seule société = un seul
            # magasin la plupart du temps, comme Smartphone.Mg) : évite
            # d'imposer un choix de magasin pour les actions qui en ont besoin
            # (caisse...) quand il n'y en a qu'un seul de toute façon.
            own_magasins = MagasinProfile.objects.filter(Q(admin=user) | Q(admins=user)).distinct()
            if own_magasins.count() == 1:
                only = own_magasins.first()
                data["magasin_id"] = only.id
                data["shop_name"] = only.shop_name
        elif user.role == "magasin":
            try:
                p = user.magasin_profile
                data["shop_name"] = p.shop_name
                data["magasin_id"] = p.id
                data["shop_logo"] = request.build_absolute_uri(p.shop_logo.url) if p.shop_logo else None
            except Exception:
                pass
        elif user.role == "employer":
            try:
                p = user.employer_profile
                data["position"] = p.position
                data["commande_role"] = p.commande_role
                if p.magasin:
                    data["magasin_id"] = p.magasin.id
                    data["shop_name"] = p.magasin.shop_name
                    data["shop_logo"] = request.build_absolute_uri(p.magasin.shop_logo.url) if p.magasin.shop_logo else None
            except Exception:
                pass
        return Response(data)

    def patch(self, request):
        user = request.user
        full_name = request.data.get("full_name")
        phone = request.data.get("phone")
        if full_name:
            user.full_name = full_name
        if phone is not None:
            user.phone = phone
        user.save()

        if user.role == "admin":
            company_name = request.data.get("company_name")
            logo = request.data.get("logo")
            try:
                p = user.admin_profile
            except AdminProfile.DoesNotExist:
                # Co-admin: no AdminProfile of their own, and none should be
                # created here — that would silently turn them into an
                # "owner" (see is_company_owner) the next time they just
                # update their name/phone.
                p = None
            if p:
                if company_name is not None:
                    p.company_name = company_name
                if logo is not None and not isinstance(logo, str):
                    p.logo = logo
                p.save()
        elif user.role == "magasin":
            shop_name = request.data.get("shop_name")
            shop_logo = request.data.get("shop_logo")
            try:
                p = user.magasin_profile
            except MagasinProfile.DoesNotExist:
                p = None
            if p:
                if shop_name is not None:
                    p.shop_name = shop_name
                if shop_logo is not None and not isinstance(shop_logo, str):
                    p.shop_logo = shop_logo
                p.save()

        return Response({"message": "Profil mis à jour"})

# =========================
# ROLE MANAGEMENT
# =========================
class RoleManagementView(APIView):
    permission_classes = [IsAuthenticated, IsAdmin]
    def put(self, request, user_id):
        current_user = request.user
        user_admin = get_user_admin(current_user)
        if not user_admin:
            return Response({"error": "Permission refusée : entreprise introuvable."}, status=403)

        new_role = request.data.get("role")
        if new_role not in ["admin", "magasin", "employer"]:
            return Response({"error": "Rôle invalide. Les rôles valides sont: admin, magasin, employer"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = CustomUser.objects.get(id=user_id)
        except CustomUser.DoesNotExist:
            return Response({"error": "Utilisateur introuvable"}, status=status.HTTP_404_NOT_FOUND)

        if user.id == current_user.id:
            return Response({"error": "Vous ne pouvez pas modifier votre propre rôle"}, status=400)

        # Touching an existing admin account, or promoting someone to admin,
        # is company-ownership territory: reserved to the founder, never to
        # a co-admin (mirrors AddAdminView) — otherwise a co-admin could
        # bypass that restriction by promoting/demoting through this endpoint.
        if user.role == "admin" or new_role == "admin":
            if not is_company_owner(current_user):
                return Response({"error": "Seul le fondateur de la société peut gérer les administrateurs."}, status=403)
            if user.role == "admin" and is_company_owner(user):
                return Response({"error": "Action impossible sur le fondateur de la société."}, status=403)

        if user.role == "admin":
            # Admin accounts have no magasin_profile/employer_profile — the
            # only way to confirm they're actually a co-admin of THIS
            # company is via the shared magasins' admins M2M.
            is_member = MagasinProfile.objects.filter(Q(admin=user_admin) | Q(admins=user_admin), admins=user).exists()
        else:
            is_member = CustomUser.objects.filter(
                Q(id=user_id),
                Q(magasin_profile__admin=user_admin) | Q(employer_profile__admin=user_admin) | Q(employer_profile__magasin__admin=user_admin)
            ).exists()
        if not is_member:
            return Response({"error": "Permission refusée : cet utilisateur n'appartient pas à votre entreprise."}, status=403)

        old_role = user.role
        user.role = new_role
        user.save()

        if new_role == "admin" and old_role != "admin":
            # Mirror AddAdminView: a freshly promoted admin must get the same
            # magasin access as the founder, otherwise they'd be an "admin"
            # with zero visible stores.
            magasins = MagasinProfile.objects.filter(Q(admin=current_user) | Q(admins=current_user)).distinct()
            for magasin in magasins:
                magasin.admins.add(user)

        return Response({
            "message": f"Rôle modifié de {old_role} à {new_role}",
            "user_id": user.id,
            "email": user.email,
            "new_role": user.role,
        })


class EmployerCommandeRoleUpdateView(APIView):
    """Assigne le sous-rôle Préparateur/Livreur du module Commande à un
    employé (§4/§5 Smartreadme.md) — réservé au gérant (admin ou magasin)."""

    permission_classes = [IsAuthenticated, IsGerant]

    def put(self, request, user_id):
        try:
            employer = EmployerProfile.objects.get(
                user_id=user_id, magasin__in=get_accessible_magasins(request.user)
            )
        except EmployerProfile.DoesNotExist:
            return Response({"error": "Employé introuvable"}, status=404)

        commande_role = request.data.get("commande_role") or None
        valid_choices = [choice[0] for choice in EmployerProfile.COMMANDE_ROLE_CHOICES]
        if commande_role is not None and commande_role not in valid_choices:
            return Response({"error": "commande_role invalide (PREPARATEUR ou LIVREUR)"}, status=400)

        employer.commande_role = commande_role
        employer.save(update_fields=["commande_role"])
        return Response({
            "message": "Rôle module Commande mis à jour",
            "user_id": user_id,
            "commande_role": commande_role,
        })


# =========================
def _accessible_magasins(user):
    """MagasinProfile queryset visible to `user` given their role — même
    règle de scoping que catalog/orders/suppliers (voir aussi
    users.permissions.get_accessible_magasins, factorisée pour ces apps), car
    réutilisée par les deux viewsets caisse ci-dessous."""
    if user.role == "admin":
        return MagasinProfile.objects.filter(Q(admin=user) | Q(admins=user)).distinct()
    if user.role == "magasin":
        return MagasinProfile.objects.filter(user=user)
    if user.role == "employer":
        return MagasinProfile.objects.filter(employers__user=user)
    return MagasinProfile.objects.none()


def _parse_custom_datetime(raw, field_label):
    """Optional custom timestamp for caisse open/close (backdating: the
    gérant records the session after the fact). Returns (value, error) —
    `value` is None (meaning "use now()") when `raw` is empty, `error` is a
    ready-to-return Response when `raw` is present but invalid."""
    if not raw:
        return None, None
    parsed = parse_datetime(raw)
    if parsed is None:
        return None, Response({"error": f"{field_label} invalide (format attendu : ISO 8601)."}, status=400)
    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed)
    if parsed > timezone.now():
        return None, Response({"error": f"{field_label} ne peut pas être dans le futur."}, status=400)
    return parsed, None


def _resolve_own_magasin(request):
    """Magasin sur lequel `request.user` peut ouvrir/alimenter une caisse :
    le sien pour magasin/employer, celui désigné par `magasin_id`/`magasin`
    (et possédé par sa société) pour un admin — qui n'a pas de magasin propre."""
    user = request.user
    if user.role == "magasin":
        return MagasinProfile.objects.filter(user=user).first()
    if user.role == "employer":
        try:
            return user.employer_profile.magasin
        except EmployerProfile.DoesNotExist:
            return None
    if user.role == "admin":
        magasin_id = request.data.get("magasin_id") or request.data.get("magasin") or request.query_params.get("magasin_id")
        if not magasin_id:
            return None
        return _accessible_magasins(user).filter(id=magasin_id).first()
    return None


class CaisseSessionViewSet(viewsets.ModelViewSet):
    """Sessions de caisse (ouverture/fermeture) — pas de `create`/`update`
    génériques : on passe par les actions `open`/`close` ci-dessous pour
    garder les calculs (solde attendu, écart) et la règle "une session
    ouverte à la fois par magasin" au même endroit."""

    serializer_class = CaisseSessionSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ["get", "post", "head", "options"]

    def get_queryset(self):
        qs = CaisseSession.objects.select_related("magasin", "opened_by", "closed_by").prefetch_related("movements")
        qs = qs.filter(magasin__in=_accessible_magasins(self.request.user))

        magasin_id = self.request.query_params.get("magasin_id") or self.request.query_params.get("store_id")
        if magasin_id:
            qs = qs.filter(magasin_id=magasin_id)
        status_param = self.request.query_params.get("status")
        if status_param:
            qs = qs.filter(status=status_param)
        return qs

    def create(self, request, *args, **kwargs):
        return Response({"error": "Utiliser POST /caisse/sessions/open/ pour ouvrir une session."}, status=405)

    @action(detail=False, methods=["get"])
    def current(self, request):
        """Session actuellement ouverte pour le magasin résolu (ou
        `magasin_id` en query pour un admin) — `204 No Content` si aucune
        (DRF ne sérialise pas `None` en `null` JSON, voir Response(None))."""
        qs = self.get_queryset().filter(status="open")
        magasin_id = request.query_params.get("magasin_id")
        if magasin_id:
            qs = qs.filter(magasin_id=magasin_id)
        elif request.user.role != "admin":
            magasin = _resolve_own_magasin(request)
            qs = qs.filter(magasin=magasin) if magasin else CaisseSession.objects.none()
        session = qs.order_by("-opened_at").first()
        if session is None:
            return Response(status=204)
        return Response(self.get_serializer(session).data)

    @action(detail=False, methods=["post"])
    def open(self, request):
        magasin = _resolve_own_magasin(request)
        if magasin is None:
            return Response({"error": "Magasin introuvable ou non spécifié."}, status=400)
        if CaisseSession.objects.filter(magasin=magasin, status="open").exists():
            return Response({"error": "Une session de caisse est déjà ouverte pour ce magasin."}, status=400)
        try:
            opening_balance = Decimal(str(request.data.get("opening_balance", 0)))
        except InvalidOperation:
            return Response({"error": "Montant d'ouverture invalide."}, status=400)
        opened_at, error = _parse_custom_datetime(request.data.get("opened_at"), "Heure d'ouverture")
        if error:
            return error

        session = CaisseSession.objects.create(
            magasin=magasin,
            opened_by=request.user,
            opening_balance=opening_balance,
            opening_note=request.data.get("opening_note") or None,
            **({"opened_at": opened_at} if opened_at else {}),
        )
        return Response(self.get_serializer(session).data, status=201)

    @action(detail=True, methods=["post"])
    def close(self, request, pk=None):
        session = self.get_object()
        if session.status == "closed":
            return Response({"error": "Cette session est déjà fermée."}, status=400)
        if request.data.get("closing_balance") is None:
            return Response({"error": "Montant de fermeture requis."}, status=400)
        try:
            closing_balance = Decimal(str(request.data.get("closing_balance")))
        except InvalidOperation:
            return Response({"error": "Montant de fermeture invalide."}, status=400)
        closed_at, error = _parse_custom_datetime(request.data.get("closed_at"), "Heure de fermeture")
        if error:
            return error
        if closed_at and closed_at < session.opened_at:
            return Response({"error": "L'heure de fermeture ne peut pas être avant l'heure d'ouverture."}, status=400)

        totals = session.movements.aggregate(
            total_in=Sum("amount", filter=Q(movement_type="in")),
            total_out=Sum("amount", filter=Q(movement_type="out")),
        )
        expected = session.opening_balance + (totals["total_in"] or 0) - (totals["total_out"] or 0)

        session.status = "closed"
        session.closed_by = request.user
        session.closed_at = closed_at or timezone.now()
        session.closing_balance = closing_balance
        session.expected_balance = expected
        session.difference = closing_balance - expected
        session.closing_note = request.data.get("closing_note") or None
        session.save()
        return Response(self.get_serializer(session).data)


class CaisseMovementViewSet(viewsets.ModelViewSet):
    serializer_class = CaisseMovementSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ["get", "post", "delete", "head", "options"]

    def get_queryset(self):
        qs = CaisseMovement.objects.select_related("magasin", "created_by", "session")
        qs = qs.filter(magasin__in=_accessible_magasins(self.request.user))

        session_id = self.request.query_params.get("session_id") or self.request.query_params.get("session")
        if session_id:
            qs = qs.filter(session_id=session_id)
        magasin_id = self.request.query_params.get("magasin_id") or self.request.query_params.get("store_id")
        if magasin_id:
            qs = qs.filter(magasin_id=magasin_id)
        return qs

    def perform_create(self, serializer):
        user = self.request.user
        session_id = self.request.data.get("session")
        if session_id:
            session = CaisseSession.objects.filter(
                id=session_id, magasin__in=_accessible_magasins(user)
            ).first()
        else:
            magasin = _resolve_own_magasin(self.request)
            session = (
                CaisseSession.objects.filter(magasin=magasin, status="open").order_by("-opened_at").first()
                if magasin
                else None
            )
        if session is None:
            raise serializers.ValidationError("Aucune session de caisse ouverte.")
        if session.status != "open":
            raise serializers.ValidationError("Cette session de caisse est fermée.")
        serializer.save(session=session, magasin=session.magasin, created_by=user)


# =========================
# NOTIFICATION VIEWSET
# =========================
class NotificationViewSet(viewsets.ModelViewSet):
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        qs = Notification.objects.select_related('magasin', 'caisse_session', 'user')
        if user.role == 'admin':
            return qs.filter(Q(magasin__admins=user) | Q(user=user)).distinct()
        elif user.role == 'magasin':
            try:
                magasin = MagasinProfile.objects.get(user=user)
                return qs.filter(Q(magasin=magasin) | Q(user=user)).distinct()
            except MagasinProfile.DoesNotExist:
                return qs.filter(user=user)
        elif user.role == 'employer':
            try:
                employer = EmployerProfile.objects.filter(user=user).first()
                if employer and employer.magasin:
                    return qs.filter(Q(magasin=employer.magasin) | Q(user=user)).distinct()
            except Exception:
                pass
            return qs.filter(user=user)

    @action(detail=False, methods=['post'], url_path='mark-all-read')
    def mark_all_read(self, request):
        """Marque toutes les notifications visibles comme lues."""
        self.get_queryset().update(is_read=True)
        return Response({"message": "Toutes les notifications marquées comme lues."})

    @action(detail=False, methods=['post'], url_path='delete-all')
    def delete_all(self, request):
        """Supprime toutes les notifications visibles."""
        self.get_queryset().delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'], url_path='bulk-delete')
    def bulk_delete(self, request):
        """Supprime une sélection spécifique de notifications."""
        ids = request.data.get('ids', [])
        self.get_queryset().filter(id__in=ids).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'], url_path='bulk-read')
    def bulk_read(self, request):
        """Marque une sélection spécifique comme lue."""
        ids = request.data.get('ids', [])
        self.get_queryset().filter(id__in=ids).update(is_read=True)
        return Response({"message": "Notifications marquées comme lues."})

    def partial_update(self, request, *args, **kwargs):
        # allow marking as read
        instance = self.get_object()
        is_read = request.data.get('is_read', None)
        if is_read is not None:
            instance.is_read = bool(is_read)
            instance.save()
            return Response(self.get_serializer(instance).data)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        user = request.user

        if user.role == 'admin':
            return super().destroy(request, *args, **kwargs)

        if instance.user_id == user.id:
            return super().destroy(request, *args, **kwargs)

        if user.role in ['magasin', 'employer']:
            magasin = None
            try:
                if user.role == 'magasin':
                    magasin = MagasinProfile.objects.get(user=user)
                else:
                    employer = EmployerProfile.objects.filter(user=user).first()
                    magasin = employer.magasin if employer else None
            except Exception:
                magasin = None
            if instance.magasin and magasin and instance.magasin.id == magasin.id:
                return super().destroy(request, *args, **kwargs)

        return Response({"error": "Permission refusée"}, status=403)

# =========================
# USERS BY MAGASIN VIEW
# =========================
class LogoutEventView(APIView):
    """Best-effort logout timestamp: JWT sessions have no server-side
    invalidation, so this only records an explicit "Déconnexion" click, not
    token expiry or the tab being closed."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        last_event = LoginEvent.objects.filter(user=request.user).order_by("-created_at").first()
        if last_event:
            last_event.logged_out_at = timezone.now()
            last_event.save(update_fields=["logged_out_at"])
        return Response({"message": "Déconnexion enregistrée"})


class UsersByMagasinView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        user = request.user
        response_data = []
        company_users = []
        seen_user_ids = set()
        device_cache = {}
        login_event_cache = {}

        def get_login_times(user_id):
            """Most recent login/logout timestamps for a user, memoized so a
            company with many members doesn't re-query per magasin."""
            if user_id not in login_event_cache:
                last_event = LoginEvent.objects.filter(user_id=user_id).order_by("-created_at").first()
                login_event_cache[user_id] = {
                    "last_login_at": last_event.created_at if last_event else None,
                    "last_logout_at": last_event.logged_out_at if last_event else None,
                }
            return login_event_cache[user_id]

        def get_devices_for_admin_profile(admin_profile):
            """Latest connected Device per user, for the given company —
            memoized so a company with several magasins doesn't re-query."""
            if not admin_profile:
                return {}
            if admin_profile.id not in device_cache:
                latest = {}
                for d in Device.objects.filter(admin_profile=admin_profile).order_by("-last_seen"):
                    if d.user_id and d.user_id not in latest:
                        latest[d.user_id] = d
                device_cache[admin_profile.id] = latest
            return device_cache[admin_profile.id]

        def build_device_info(device):
            if not device:
                return None
            return {
                "device_name": parse_device_name(device.user_agent),
                "ip_address": device.ip_address,
                "last_seen": device.last_seen,
                "latitude": device.latitude,
                "longitude": device.longitude,
            }

        def add_company_user(user_obj, shop_name=None, magasin_id=None, position=None, device=None):
            if not user_obj or user_obj.id in seen_user_ids:
                return
            seen_user_ids.add(user_obj.id)
            company_users.append({
                "id": user_obj.id,
                "full_name": user_obj.full_name,
                "email": user_obj.email,
                "is_confirmed": user_obj.is_confirmed,
                "role": user_obj.role,
                "shop_name": shop_name,
                "magasin_id": magasin_id,
                "position": position,
                "device": build_device_info(device),
                **get_login_times(user_obj.id),
            })

        # Admin can see all magasins belonging to them
        if user.role == "admin":
            magasins = MagasinProfile.objects.filter(Q(admin=user) | Q(admins=user)).distinct()
        # Magasin can see only their own magasin
        elif user.role == "magasin":
            magasins = MagasinProfile.objects.filter(user=user)
        # Employer can see only their own magasin
        elif user.role == "employer":
            try:
                employer_profile = EmployerProfile.objects.get(user=user)
                if employer_profile.magasin:
                    magasins = MagasinProfile.objects.filter(id=employer_profile.magasin.id)
                else:
                    return Response([])
            except EmployerProfile.DoesNotExist:
                return Response({"error": "Employer profile not found"}, status=404)
        else:
            return Response({"error": "Role not supported"}, status=403)

        for mag in magasins:
            admin_profile = getattr(mag.admin, "admin_profile", None) if mag.admin else None
            devices_by_user = get_devices_for_admin_profile(admin_profile)

            manager_data = {
                "id": mag.admin.id,
                "full_name": mag.admin.full_name,
                "email": mag.admin.email,
                "is_confirmed": mag.admin.is_confirmed,
                "role": mag.admin.role,
                "device": build_device_info(devices_by_user.get(mag.admin.id)),
                **get_login_times(mag.admin.id),
            } if mag.admin else None
            if manager_data:
                add_company_user(mag.admin, mag.shop_name, mag.id, device=devices_by_user.get(mag.admin.id))

            # Le vrai compte "gérant" (role=magasin, MagasinProfile.user) n'était
            # jamais exposé nulle part dans cette réponse — seul l'admin de la
            # société (mag.admin, ci-dessus) apparaissait sous le nom "manager".
            # Un gérant nouvellement créé/approuvé restait donc invisible dans
            # toute liste basée sur cet endpoint (magasins/users/).
            if mag.user:
                add_company_user(mag.user, mag.shop_name, mag.id, device=devices_by_user.get(mag.user.id))

            employers_qs = EmployerProfile.objects.filter(magasin=mag)
            employers_list = []
            for emp in employers_qs:
                employers_list.append({
                    "id": emp.user.id,
                    "full_name": emp.user.full_name,
                    "email": emp.user.email,
                    "is_confirmed": emp.user.is_confirmed,
                    "position": emp.position,
                    "role": emp.user.role,
                    "commande_role": emp.commande_role,
                    "device": build_device_info(devices_by_user.get(emp.user.id)),
                    **get_login_times(emp.user.id),
                })
                add_company_user(emp.user, mag.shop_name, mag.id, emp.position, device=devices_by_user.get(emp.user.id))

            for admin_user in mag.admins.all():
                add_company_user(admin_user, mag.shop_name, mag.id, device=devices_by_user.get(admin_user.id))

            response_data.append({
                "magasin_id": mag.id,
                "shop_name": mag.shop_name,
                "shop_logo": request.build_absolute_uri(mag.shop_logo.url) if mag.shop_logo else None,
                "manager": manager_data,
                "employers": employers_list,
                "company_users": company_users,
            })

        return Response(response_data)


class MagasinStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        if user.role == "admin":
            magasins = MagasinProfile.objects.filter(Q(admin=user) | Q(admins=user)).distinct()
        elif user.role == "magasin":
            magasins = MagasinProfile.objects.filter(user=user)
        elif user.role == "employer":
            try:
                employer_profile = EmployerProfile.objects.get(user=user)
                if employer_profile.magasin:
                    magasins = MagasinProfile.objects.filter(id=employer_profile.magasin.id)
                else:
                    return Response([])
            except EmployerProfile.DoesNotExist:
                return Response({"error": "Employer profile not found"}, status=404)
        else:
            return Response({"error": "Role not supported"}, status=403)

        response_data = []

        for mag in magasins:
            magasin_qs = MagasinProfile.objects.filter(id=mag.id)
            variants_qs = ProductVariant.objects.filter(product_reference__type__category__magasin=mag)
            livrees_qs = Order.objects.filter(magasin=mag, statut_courant="LIVRE")

            total_products = ProductReference.objects.filter(type__category__magasin=mag).count()
            total_stock_quantity = variants_qs.aggregate(total=Coalesce(Sum('stock_actuel'), 0))['total']
            total_stock_value = _stock_value_for_magasins(magasin_qs)
            total_sold_value = livrees_qs.aggregate(
                total=Coalesce(Sum('total_a_payer', output_field=DecimalField()), 0, output_field=DecimalField())
            )['total']
            cost_map = _cost_by_variant(magasin_qs)
            _, profit = _order_items_revenue_and_profit(livrees_qs, cost_map)

            response_data.append({
                "magasin_id": mag.id,
                "shop_name": mag.shop_name,
                "total_products": total_products,
                "total_stock_quantity": total_stock_quantity,
                "total_stock_value": total_stock_value,
                "total_sold_value": total_sold_value,
                "profit": profit,
            })

        return Response(response_data)


# =========================
# DASHBOARD VIEW
# =========================
class DashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        role = user.role
        today = timezone.now().date()

        if role == "admin":
            # KPIs scoped to admin (owner via FK, or co-admin via the admins M2M)
            admin_magasins = MagasinProfile.objects.filter(Q(admin=user) | Q(admins=user)).distinct()
            admin_magasin_ids = list(admin_magasins.values_list("id", flat=True))
            admin_orders = Order.objects.filter(magasin_id__in=admin_magasin_ids)
            admin_livrees = admin_orders.filter(statut_courant="LIVRE")
            admin_variants = ProductVariant.objects.filter(
                product_reference__type__category__magasin_id__in=admin_magasin_ids
            )
            admin_employers = EmployerProfile.objects.filter(
                Q(admin=user) | Q(magasin__admin=user) | Q(magasin__admins=user)
            ).distinct()

            cost_map = _cost_by_variant(admin_magasins)
            total_revenue, total_profit = _order_items_revenue_and_profit(admin_livrees, cost_map)
            today_livrees = admin_livrees.filter(updated_at__date=today)
            _, profit_today = _order_items_revenue_and_profit(today_livrees, cost_map)

            total_stock_value = _stock_value_for_magasins(admin_magasins)
            total_magasins = admin_magasins.count()
            total_employers = admin_employers.count()
            total_products = ProductReference.objects.filter(type__category__magasin_id__in=admin_magasin_ids).count()
            total_sales = admin_livrees.count()
            sales_today = today_livrees.count()
            low_stock_count = admin_variants.filter(stock_actuel__lte=F('seuil_alerte')).count()

            # Le catalogue Smartreadme.md (housses/cache-écran) ne suit pas de
            # date de péremption — champs gardés à 0 pour la stabilité des
            # cartes dashboard existantes plutôt que supprimés (§8.1).
            expired_count = 0
            expiring_soon_count = 0
            unpaid_sales_count, unpaid_sales_value = 0, 0  # paiement partiel hors périmètre MVP (§3)

            top_products, bottom_products = _top_bottom_products(admin_magasin_ids, admin_livrees)

            low_stock_list = [
                {
                    "name": f"{v.product_reference.reference_name} ({v.couleur})",
                    "initial_quantity": v.stock_actuel,
                    "alert_threshold": v.seuil_alerte,
                    "magasin__shop_name": v.product_reference.type.category.magasin.shop_name,
                }
                for v in admin_variants.filter(stock_actuel__lte=F('seuil_alerte')).select_related(
                    "product_reference", "product_reference__type__category__magasin"
                )[:5]
            ]
            expired_list = []
            expiring_soon_list = []

            recent_orders_qs = admin_livrees.select_related("magasin", "created_by").prefetch_related(
                "items__product_variant__product_reference"
            ).order_by('-updated_at')[:5]
            recent_sales = [
                {
                    "product_name": ", ".join(
                        f"{item.product_variant.product_reference.reference_name} x{item.quantite}"
                        for item in order.items.all()
                    ),
                    "quantity": sum(item.quantite for item in order.items.all()),
                    "sale_price": None,
                    "total_price": order.total_a_payer,
                    "seller_name": order.created_by.full_name if order.created_by else None,
                    "shop_name": order.magasin.shop_name if order.magasin else None,
                    "sold_at": order.updated_at,
                }
                for order in recent_orders_qs
            ]

            best_employees = list(
                admin_livrees.values('created_by__full_name')
                .annotate(sales_count=Count('id'), total_amount=Sum('total_a_payer'))
                .order_by('-total_amount')[:5]
            )
            best_shops = list(
                admin_livrees.values('magasin__shop_name')
                .annotate(total_amount=Sum('total_a_payer'), sales_count=Count('id'))
                .order_by('-total_amount')[:5]
            )
            for row in best_shops:
                row["total_stock"] = 0

            return Response({
                "role": role,
                "kpis": {
                    "total_revenue": total_revenue,
                    "total_profit": total_profit,
                    "total_stock_value": total_stock_value,
                    "total_magasins": total_magasins,
                    "total_employers": total_employers,
                    "total_products": total_products,
                    "total_sales": total_sales,
                    "sales_today": sales_today,
                    "profit_today": profit_today,
                    "low_stock_count": low_stock_count,
                    "expired_count": expired_count,
                    "expiring_soon_count": expiring_soon_count,
                    "unpaid_sales_count": unpaid_sales_count,
                    "unpaid_sales_value": unpaid_sales_value,
                },
                "lists": {
                    "top_products": top_products,
                    "bottom_products": bottom_products,
                    "low_stock_products": low_stock_list,
                    "expired_products": expired_list,
                    "expiring_soon_products": expiring_soon_list,
                    "recent_sales": recent_sales,
                    "best_employees": best_employees,
                    "best_shops": best_shops
                }
            })

        elif role == "magasin":
            try:
                magasin = MagasinProfile.objects.get(user=user)
            except MagasinProfile.DoesNotExist:
                return Response({"error": "Magasin profile not found"}, status=404)

            magasin_qs = MagasinProfile.objects.filter(id=magasin.id)
            orders_qs = Order.objects.filter(magasin=magasin)
            livrees_qs = orders_qs.filter(statut_courant="LIVRE")
            variants_qs = ProductVariant.objects.filter(product_reference__type__category__magasin=magasin)

            cost_map = _cost_by_variant(magasin_qs)
            total_revenue, total_profit = _order_items_revenue_and_profit(livrees_qs, cost_map)
            today_livrees = livrees_qs.filter(updated_at__date=today)
            _, profit_today = _order_items_revenue_and_profit(today_livrees, cost_map)
            sales_today = today_livrees.count()

            stock_value = _stock_value_for_magasins(magasin_qs)
            total_products = ProductReference.objects.filter(type__category__magasin=magasin).count()
            total_sales = livrees_qs.count()
            low_stock_count = variants_qs.filter(stock_actuel__lte=F('seuil_alerte')).count()
            expired_count = 0  # pas de péremption sur ce catalogue (§8.1)
            unpaid_sales_count, unpaid_sales_value = 0, 0  # hors périmètre MVP (§3)

            top_products, bottom_products = _top_bottom_products([magasin.id], livrees_qs)

            low_stock_list = [
                {"name": f"{v.product_reference.reference_name} ({v.couleur})", "initial_quantity": v.stock_actuel}
                for v in variants_qs.filter(stock_actuel__lte=F('seuil_alerte')).select_related("product_reference")[:5]
            ]

            recent_orders_qs = livrees_qs.select_related("created_by").prefetch_related(
                "items__product_variant__product_reference"
            ).order_by('-updated_at')[:5]
            recent_sales = [
                {
                    "product_name": ", ".join(
                        f"{item.product_variant.product_reference.reference_name} x{item.quantite}"
                        for item in order.items.all()
                    ),
                    "quantity": sum(item.quantite for item in order.items.all()),
                    "total_price": order.total_a_payer,
                    "seller_name": order.created_by.full_name if order.created_by else None,
                    "sold_at": order.updated_at,
                }
                for order in recent_orders_qs
            ]

            best_sellers = list(
                livrees_qs.values('created_by__full_name')
                .annotate(sales_count=Count('id'), total_amount=Sum('total_a_payer'))
                .order_by('-total_amount')[:5]
            )

            return Response({
                "role": role,
                "kpis": {
                    "sales_today": sales_today,
                    "profit_today": profit_today,
                    "total_revenue": total_revenue,
                    "total_profit": total_profit,
                    "stock_value": stock_value,
                    "total_products": total_products,
                    "total_sales": total_sales,
                    "low_stock_count": low_stock_count,
                    "expired_count": expired_count,
                    "unpaid_sales_count": unpaid_sales_count,
                    "unpaid_sales_value": unpaid_sales_value,
                },
                "lists": {
                    "top_products": top_products,
                    "bottom_products": bottom_products,
                    "low_stock_products": low_stock_list,
                    "recent_sales": recent_sales,
                    "best_sellers": best_sellers
                }
            })

        elif role == "employer":
            try:
                employer_profile = EmployerProfile.objects.get(user=user)
                magasin = employer_profile.magasin
            except EmployerProfile.DoesNotExist:
                return Response({"error": "Employer profile not found"}, status=404)

            # Un employer n'a plus de "ventes" propres (module Commande) — son
            # activité se lit dans OrderStatusHistory : le statut qu'il fait
            # avancer dépend de son commande_role (§7.2/§7.3 Smartreadme.md).
            commande_role = employer_profile.commande_role
            target_status = "LIVRE" if commande_role == "LIVREUR" else "PRETE"

            my_actions_qs = OrderStatusHistory.objects.filter(changed_by=user, nouveau_statut=target_status)
            my_orders_ids_today = my_actions_qs.filter(timestamp__date=today).values_list("order_id", flat=True)
            my_orders_ids_all = my_actions_qs.values_list("order_id", flat=True)

            orders_today_qs = Order.objects.filter(id__in=my_orders_ids_today)
            orders_all_qs = Order.objects.filter(id__in=my_orders_ids_all)

            my_sales_today = orders_today_qs.count()
            total_amount_sold = orders_all_qs.aggregate(total=Sum('total_a_payer'))['total'] or 0
            products_sold_count = OrderItem.objects.filter(order__in=orders_all_qs).aggregate(
                total=Sum('quantite')
            )['total'] or 0
            clients_count = orders_all_qs.values('client_nom').distinct().count()
            unpaid_sales_count, unpaid_sales_value = 0, 0  # hors périmètre MVP (§3)

            recent_orders_qs = orders_all_qs.prefetch_related(
                "items__product_variant__product_reference"
            ).order_by('-updated_at')[:5]
            recent_sales = [
                {
                    "product_name": ", ".join(
                        f"{item.product_variant.product_reference.reference_name} x{item.quantite}"
                        for item in order.items.all()
                    ),
                    "quantity": sum(item.quantite for item in order.items.all()),
                    "total_price": order.total_a_payer,
                    "sold_at": order.updated_at,
                }
                for order in recent_orders_qs
            ]

            return Response({
                "role": role,
                "commande_role": commande_role,
                "kpis": {
                    "my_sales_today": my_sales_today,
                    "total_amount_sold": total_amount_sold,
                    "products_sold_count": products_sold_count,
                    "clients_count": clients_count,
                    "unpaid_sales_count": unpaid_sales_count,
                    "unpaid_sales_value": unpaid_sales_value,
                },
                "lists": {
                    "recent_sales": recent_sales
                }
            })

        else:
            return Response({"error": "Role not supported"}, status=403)


# =========================
# API ENDPOINTS LIST VIEW
# =========================
class ApiEndpointsListView(APIView):
    permission_classes = []  # Publicly accessible endpoint exploration

    def get(self, request):
        endpoints = [
            {
                "path": "/api/users/login/",
                "method": "POST",
                "auth_required": False,
                "roles_allowed": ["Any"],
                "description": "Authentifie un utilisateur et retourne les tokens JWT (access & refresh)."
            },
            {
                "path": "/api/users/refresh/",
                "method": "POST",
                "auth_required": False,
                "roles_allowed": ["Any"],
                "description": "Rafraîchit le token d'accès JWT expiré."
            },
            {
                "path": "/api/users/register/",
                "method": "POST",
                "auth_required": False,
                "roles_allowed": ["Any"],
                "description": "Inscrit un nouvel utilisateur (admin créé automatiquement, magasin/employé en attente)."
            },
            {
                "path": "/api/users/me/",
                "method": "GET",
                "auth_required": True,
                "roles_allowed": ["admin", "magasin", "employer"],
                "description": "Retourne le profil complet et les informations de l'utilisateur connecté."
            },
            {
                "path": "/api/users/approve/<user_id>/",
                "method": "PUT",
                "auth_required": True,
                "roles_allowed": ["admin", "magasin"],
                "description": "Approuve et active un compte utilisateur en attente de validation."
            },
            {
                "path": "/api/users/role/<user_id>/",
                "method": "PUT",
                "auth_required": True,
                "roles_allowed": ["admin"],
                "description": "Modifie le rôle d'un utilisateur existant."
            },
            {
                "path": "/api/users/products/",
                "method": "GET, POST",
                "auth_required": True,
                "roles_allowed": ["admin", "magasin", "employer"],
                "description": "GET: Liste les produits (prix d'achat masqué pour magasin/employé). POST: Crée un nouveau produit."
            },
            {
                "path": "/api/users/products/<id>/",
                "method": "GET, PUT, PATCH, DELETE",
                "auth_required": True,
                "roles_allowed": ["admin", "magasin", "employer"],
                "description": "Consulte, modifie ou supprime un produit spécifique (Modifications/Suppression réservées aux admins)."
            },
            {
                "path": "/api/users/sales/",
                "method": "GET, POST",
                "auth_required": True,
                "roles_allowed": ["admin", "magasin", "employer"],
                "description": "GET: Historique des ventes de produits (filtré par magasin). POST: Enregistre une nouvelle transaction de vente."
            },
            {
                "path": "/api/users/sales/totals/",
                "method": "GET",
                "auth_required": True,
                "roles_allowed": ["admin", "magasin", "employer"],
                "description": "Calcule la somme globale des unit_price et shell_price de tous les produits."
            },
            {
                "path": "/api/users/sales/profit/",
                "method": "GET",
                "auth_required": True,
                "roles_allowed": ["admin"],
                "description": "Calcule le bénéfice réel total (somme de (sale_price - unit_price) * quantity) pour l'administrateur."
            },
            {
                "path": "/api/users/sales/profit-by-magasins/",
                "method": "GET",
                "auth_required": True,
                "roles_allowed": ["admin"],
                "description": "Liste le bénéfice total, le chiffre d'affaires et le coût pour chaque magasin appartenant à l'administrateur."
            },
            {
                "path": "/api/users/magasins/users/",
                "method": "GET",
                "auth_required": True,
                "roles_allowed": ["admin", "magasin", "employer"],
                "description": "Retourne la liste de tous les utilisateurs (managers et employés) regroupés par magasin."
            },
            {
                "path": "/api/users/dashboard/",
                "method": "GET",
                "auth_required": True,
                "roles_allowed": ["admin", "magasin", "employer"],
                "description": "Tableau de bord analytique dynamique adapté en temps réel au profil de l'utilisateur."
            },
            {
                "path": "/api/users/endpoints/",
                "method": "GET",
                "auth_required": False,
                "roles_allowed": ["Any"],
                "description": "Liste l'ensemble des endpoints disponibles avec leurs descriptions et permissions."
            }
        ]
        return Response(endpoints)


# =========================
# CHANGE PASSWORD VIEW
# =========================
class ChangePasswordView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        old_password = request.data.get("old_password")
        new_password = request.data.get("new_password")
        if not old_password or not new_password:
            return Response({"detail": "Champs requis manquants"}, status=400)
        if not user.check_password(old_password):
            return Response({"detail": "Mot de passe actuel incorrect"}, status=400)
        if len(new_password) < 6:
            return Response({"detail": "Le mot de passe doit contenir au moins 6 caractères"}, status=400)
        user.set_password(new_password)
        user.save()
        return Response({"message": "Mot de passe changé avec succès"})


# =========================
# PENDING USERS VIEW
# =========================
class PendingUsersView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if user.role not in ["admin", "magasin"]:
            return Response({"error": "Permission refusée"}, status=403)

        if user.role == "admin":
            # NB: le filtre ne couvrait que les employés (employer_profile) ;
            # les gérants de magasin en attente (magasin_profile) n'apparaissaient
            # donc jamais ici, alors que le code plus bas sait déjà les afficher
            # (voir `elif u.role == "magasin"`) — bug d'oubli, pas un choix voulu.
            pending_qs = CustomUser.objects.filter(
            Q(is_confirmed=False),
            Q(employer_profile__admin=user)
            | Q(employer_profile__magasin__admin=user)
            | Q(magasin_profile__admin=user)
        ).distinct()
        else:
            try:
                magasin = MagasinProfile.objects.get(user=user)
                employer_ids = EmployerProfile.objects.filter(magasin=magasin).values_list("user_id", flat=True)
                pending_qs = CustomUser.objects.filter(is_confirmed=False, id__in=employer_ids)
            except MagasinProfile.DoesNotExist:
                return Response({"error": "Magasin introuvable"}, status=404)

        data = []
        for u in pending_qs:
            item = {
                "id": u.id,
                "full_name": u.full_name,
                "email": u.email,
                "role": u.role,
                "created_at": u.created_at,
            }
            if u.role == "employer":
                try:
                    ep = u.employer_profile
                    item["position"] = ep.position
                    if ep.magasin:
                        item["shop_name"] = ep.magasin.shop_name
                except Exception:
                    pass
            elif u.role == "magasin":
                try:
                    mp = u.magasin_profile
                    item["shop_name"] = mp.shop_name
                except Exception:
                    pass
            data.append(item)

        return Response(data)


# =========================
# DELETE USER VIEW
# =========================
class DeleteUserView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, user_id):
        current_user = request.user
        if current_user.role not in ["admin", "magasin"]:
            return Response({"error": "Permission refusée"}, status=403)

        password = request.data.get("password")
        if not password:
            return Response({"error": "Mot de passe requis pour confirmer la suppression."}, status=400)
        if not current_user.check_password(password):
            return Response({"error": "Mot de passe incorrect."}, status=400)

        user_admin = get_user_admin(current_user)
        if not user_admin:
            return Response({"error": "Permission refusée : entreprise introuvable."}, status=403)

        try:
            user = CustomUser.objects.get(id=user_id)
        except CustomUser.DoesNotExist:
            return Response({"error": "Utilisateur introuvable"}, status=404)

        if user.id == request.user.id:
            return Response({"error": "Vous ne pouvez pas vous supprimer vous-même"}, status=400)

        if current_user.role == "magasin":
            try:
                magasin = current_user.magasin_profile
                is_member = CustomUser.objects.filter(
                    id=user_id,
                    role="employer",
                    employer_profile__magasin=magasin
                ).exists()
                if not is_member:
                    return Response({"error": "Permission refusée : cet employé n'appartient pas à votre magasin."}, status=403)
            except Exception:
                return Response({"error": "Magasin introuvable"}, status=404)
        else: # admin
            if user.role == "admin":
                # Removing a co-admin (or the founder) is company-ownership
                # territory: reserved to the founder, and never targetable
                # at the founder themselves (mirrors RoleManagementView).
                if not is_company_owner(current_user):
                    return Response({"error": "Seul le fondateur de la société peut retirer un administrateur."}, status=403)
                if is_company_owner(user):
                    return Response({"error": "Action impossible sur le fondateur de la société."}, status=403)
                is_member = MagasinProfile.objects.filter(Q(admin=user_admin) | Q(admins=user_admin), admins=user).exists()
            else:
                is_member = CustomUser.objects.filter(
                    Q(id=user_id),
                    Q(magasin_profile__admin=user_admin) | Q(employer_profile__admin=user_admin) | Q(employer_profile__magasin__admin=user_admin)
                ).exists()
            if not is_member:
                return Response({"error": "Permission refusée : cet utilisateur n'appartient pas à votre entreprise."}, status=403)

        user.delete()
        return Response({"message": "Utilisateur supprimé"})


# =========================
# REJECT USER VIEW
# =========================
class RejectUserView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, user_id):
        current_user = request.user
        if current_user.role not in ["admin", "magasin"]:
            return Response({"error": "Permission refusée"}, status=403)

        user_admin = get_user_admin(current_user)
        if not user_admin:
            return Response({"error": "Permission refusée : entreprise introuvable."}, status=403)

        if current_user.role == "magasin":
            try:
                magasin = current_user.magasin_profile
                is_member = CustomUser.objects.filter(
                    id=user_id,
                    role="employer",
                    employer_profile__magasin=magasin
                ).exists()
                if not is_member:
                    return Response({"error": "Permission refusée : cet employé n'appartient pas à votre magasin."}, status=403)
            except Exception:
                return Response({"error": "Magasin introuvable"}, status=404)
        else: # admin
            is_member = CustomUser.objects.filter(
                Q(id=user_id),
                Q(magasin_profile__admin=user_admin) | Q(employer_profile__admin=user_admin) | Q(employer_profile__magasin__admin=user_admin)
            ).exists()
            if not is_member:
                return Response({"error": "Permission refusée : cet utilisateur n'appartient pas à votre entreprise."}, status=403)

        try:
            user = CustomUser.objects.get(id=user_id)
            user.delete()
            return Response({"message": "Utilisateur rejeté et supprimé"})
        except CustomUser.DoesNotExist:
            return Response({"error": "Utilisateur introuvable"}, status=404)


# =========================
# MAGASIN VIEWSET
# =========================
class MagasinViewSet(viewsets.ModelViewSet):
    serializer_class = MagasinProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role == "admin":
            return MagasinProfile.objects.filter(Q(admin=user) | Q(admins=user)).distinct()
        elif user.role == "magasin":
            return MagasinProfile.objects.filter(user=user)
        return MagasinProfile.objects.none()

    def perform_create(self, serializer):
        if self.request.user.role != "admin":
            raise serializers.ValidationError("Seul l'admin peut créer un magasin.")
        magasin = serializer.save(admin=self.request.user)
        # Le créateur peut être un co-admin : partager l'accès avec tous les
        # admins de la société (fondateur inclus), sinon ce magasin reste
        # invisible pour eux. Cf. RegisterSerializer.create (role="magasin").
        from .subscriptions import get_company_admin_ids
        magasin.admins.set(CustomUser.objects.filter(id__in=get_company_admin_ids(self.request.user)))

    def destroy(self, request, *args, **kwargs):
        password = request.data.get("password")
        if not password:
            return Response({"error": "Mot de passe requis pour confirmer la suppression."}, status=400)
        if not request.user.check_password(password):
            return Response({"error": "Mot de passe incorrect."}, status=400)
        return super().destroy(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        shop_name = request.data.get("shop_name")
        shop_logo = request.data.get("shop_logo")
        manager_id = request.data.get("manager_id")
        if shop_name is not None:
            instance.shop_name = shop_name
        if shop_logo is not None and not isinstance(shop_logo, str):
            instance.shop_logo = shop_logo
        if manager_id is not None:
            if request.user.role != "admin":
                return Response({"error": "Seul un administrateur peut modifier le gérant."}, status=403)
            try:
                new_manager = CustomUser.objects.get(id=manager_id, role="magasin")
            except CustomUser.DoesNotExist:
                return Response({"manager_id": "Gérant introuvable."}, status=404)
            other_profile = MagasinProfile.objects.filter(user=new_manager).exclude(id=instance.id).first()
            if other_profile:
                return Response(
                    {"manager_id": f"Ce compte est déjà gérant de \"{other_profile.shop_name}\"."},
                    status=400,
                )
            instance.user = new_manager
        instance.save()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)


# =========================
# CHAT VIEWS
# =========================
class ChatUsersListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        magasins = get_company_magasins(request.user)
        if not magasins.exists():
            return Response([])

        users = CustomUser.objects.filter(
            Q(is_confirmed=True),
            Q(magasins__in=magasins) |
            Q(admin_magasin_profiles__in=magasins) |
            Q(magasin_profile__in=magasins) |
            Q(employer_profile__magasin__in=magasins)
        ).exclude(id=request.user.id).distinct()

        data = []
        for u in users:
            user_info = {
                "id": u.id,
                "full_name": u.full_name,
                "email": u.email,
                "role": u.role,
            }
            if u.role == "magasin":
                try:
                    user_info["shop_name"] = u.magasin_profile.shop_name
                except Exception:
                    pass
            elif u.role == "employer":
                try:
                    user_info["shop_name"] = u.employer_profile.magasin.shop_name
                except Exception:
                    pass
            data.append(user_info)
        return Response(data)


class ChatMessageHistoryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.db.models import Q
        room_name = request.query_params.get("room_name", "general")
        recipient_id = request.query_params.get("recipient_id")

        my_magasins = get_company_magasins(request.user)
        if not my_magasins.exists():
            return Response([])

        if recipient_id:
            try:
                recipient = CustomUser.objects.get(id=recipient_id)
                # Verify recipient belongs to the same company (shares at least one magasin)
                recipient_magasins = get_company_magasins(recipient)
                if not recipient_magasins.exists() or not my_magasins.filter(id__in=recipient_magasins).exists():
                    return Response({"error": "Permission refusée"}, status=403)

                messages = ChatMessage.objects.filter(
                    Q(sender=request.user, recipient=recipient) |
                    Q(sender=recipient, recipient=request.user)
                ).order_by("timestamp")
            except CustomUser.DoesNotExist:
                return Response({"error": "Destinataire introuvable"}, status=404)
        else:
            company_id = get_company_id(request.user)
            if not company_id:
                return Response([])
            scoped_room = f"general_{company_id}"
            messages = ChatMessage.objects.filter(room_name=scoped_room, recipient__isnull=True).order_by("timestamp")

        # Take last 100 messages
        total_count = messages.count()
        messages = messages.select_related("sender", "recipient")[max(0, total_count - 100):]
        serializer = ChatMessageSerializer(messages, many=True)
        return Response(serializer.data)

class _TransferValidationError(Exception):
    def __init__(self, message, status=400):
        self.message = message
        self.status = status
        super().__init__(message)


class TransferProductsView(APIView):
    """Transfert de stock entre magasins d'une même société — réservé à
    l'admin (seul rôle avec visibilité sur plusieurs magasins). Le catalogue
    (Catégorie/Type/Marque) est scopé par magasin (§8.1 Smartreadme.md) :
    une variante n'a pas d'équivalent direct dans le magasin destination, on
    retrouve/crée donc la même chaîne Catégorie→Type→Marque→Référence par
    nom avant de déplacer le stock au niveau de la variante (couleur)."""

    permission_classes = [IsAuthenticated]

    def _normalize_items(self, request):
        items = request.data.get("items")
        if isinstance(items, list) and len(items) > 0:
            return items
        variant_ids = request.data.get("variant_ids", [])
        if isinstance(variant_ids, list) and len(variant_ids) > 0:
            return [{"variant_id": vid, "quantity": None} for vid in variant_ids]
        return []

    def _find_or_create_dest_variant(self, source_variant, dest_magasin):
        source_ref = source_variant.product_reference
        source_type = source_ref.type
        source_category = source_type.category
        source_brand = source_ref.brand

        dest_category, _ = ProductCategory.objects.get_or_create(
            magasin=dest_magasin, nom=source_category.nom, defaults={"ordre": source_category.ordre}
        )
        dest_type, _ = ProductType.objects.get_or_create(category=dest_category, nom=source_type.nom)
        dest_brand, _ = Brand.objects.get_or_create(magasin=dest_magasin, nom=source_brand.nom)
        dest_reference, _ = ProductReference.objects.get_or_create(
            type=dest_type, brand=dest_brand, reference_name=source_ref.reference_name,
            defaults={"prix_vente": source_ref.prix_vente},
        )
        dest_variant, _ = ProductVariant.objects.get_or_create(
            product_reference=dest_reference, couleur=source_variant.couleur,
            defaults={"seuil_alerte": source_variant.seuil_alerte},
        )
        return dest_variant

    def post(self, request):
        user = request.user
        if user.role != "admin":
            return Response({"error": "Permission refusée"}, status=403)
        source_id = request.data.get("source_magasin_id")
        destination_id = request.data.get("destination_magasin_id")
        raw_items = self._normalize_items(request)
        if not source_id or not destination_id or not raw_items:
            return Response({"error": "Paramètres manquants ou invalides"}, status=400)
        if str(source_id) == str(destination_id):
            return Response({"error": "Le magasin source et destination doivent être différents"}, status=400)

        try:
            source_magasin = MagasinProfile.objects.filter(
                Q(admin=user) | Q(admins=user), id=source_id
            ).distinct().get()
            dest_magasin = MagasinProfile.objects.filter(
                Q(admin=user) | Q(admins=user), id=destination_id
            ).distinct().get()
        except MagasinProfile.DoesNotExist:
            return Response({"error": "Magasin source ou destination introuvable ou non autorisé"}, status=404)

        transfer_note = f"Transfert du magasin {source_magasin.id} au magasin {dest_magasin.id} par {user.full_name}"
        transferred_summary = []

        try:
            with transaction.atomic():
                for raw_item in raw_items:
                    variant_id = raw_item.get("variant_id")
                    if not variant_id:
                        raise _TransferValidationError("Identifiant de variante manquant")
                    try:
                        variant = ProductVariant.objects.select_related("product_reference").get(
                            id=variant_id, product_reference__type__category__magasin=source_magasin
                        )
                    except ProductVariant.DoesNotExist:
                        raise _TransferValidationError("Certaines variantes n'appartiennent pas au magasin source")

                    requested_qty = raw_item.get("quantity")
                    available = variant.stock_actuel
                    quantity = available if requested_qty is None else int(requested_qty)
                    label = f"{variant.product_reference.reference_name} ({variant.couleur})"

                    if quantity <= 0:
                        raise _TransferValidationError(f"Quantité invalide pour {label}")
                    if quantity > available:
                        raise _TransferValidationError(f"Stock insuffisant pour {label}. Disponible : {available}.")

                    dest_variant = self._find_or_create_dest_variant(variant, dest_magasin)

                    apply_stock_movement(
                        product_variant=variant, movement_type="SORTIE", quantite=quantity,
                        origine="AJUSTEMENT", user=user, reference=transfer_note[:100], note="Transfert sortant",
                    )
                    apply_stock_movement(
                        product_variant=dest_variant, movement_type="ENTREE", quantite=quantity,
                        origine="AJUSTEMENT", user=user, reference=transfer_note[:100], note="Transfert entrant",
                    )

                    transferred_summary.append(f"{label} x{quantity}")
        except _TransferValidationError as exc:
            return Response({"error": exc.message}, status=exc.status)

        if transferred_summary:
            preview = ", ".join(transferred_summary[:3])
            if len(transferred_summary) > 3:
                preview += f" (+{len(transferred_summary) - 3} autre(s))"

            Notification.objects.create(
                notif_type="other",
                message=f"Transfert sortant vers {dest_magasin.shop_name} : {preview} — par {user.full_name}",
                magasin=source_magasin,
                user=user,
            )
            Notification.objects.create(
                notif_type="other",
                message=f"Transfert entrant depuis {source_magasin.shop_name} : {preview} — par {user.full_name}",
                magasin=dest_magasin,
                user=user,
            )

        return Response({"message": "Transfert effectué avec succès"})


class BackupExportView(APIView):
    """Export complet de la base de données (fixture JSON) + fichiers media (images, QR codes) dans un zip."""
    permission_classes = [IsAdmin]

    def get(self, request):
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            data_buffer = io.StringIO()
            call_command(
                "dumpdata",
                exclude=[
                    "contenttypes",
                    "auth.permission",
                    "admin.logentry",
                    "sessions.session",
                ],
                indent=2,
                stdout=data_buffer,
            )
            zf.writestr("data.json", data_buffer.getvalue())

            media_root = str(settings.MEDIA_ROOT)
            if os.path.isdir(media_root):
                for root, _dirs, files in os.walk(media_root):
                    for filename in files:
                        file_path = os.path.join(root, filename)
                        arcname = os.path.join("media", os.path.relpath(file_path, media_root))
                        zf.write(file_path, arcname)

        buffer.seek(0)
        filename = f"backup_{timezone.now().strftime('%Y%m%d_%H%M%S')}.zip"
        response = HttpResponse(buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class BackupImportView(APIView):
    """Restauration complète depuis un backup.zip : remplace toutes les données et les fichiers media."""
    permission_classes = [IsAdmin]
    parser_classes = [MultiPartParser]

    def post(self, request):
        uploaded = request.FILES.get("file")
        if not uploaded:
            return Response({"detail": "Aucun fichier fourni."}, status=status.HTTP_400_BAD_REQUEST)
        if not uploaded.name.lower().endswith(".zip"):
            return Response({"detail": "Le fichier doit être une archive .zip."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            zf = zipfile.ZipFile(uploaded)
        except zipfile.BadZipFile:
            return Response({"detail": "Fichier zip invalide."}, status=status.HTTP_400_BAD_REQUEST)

        with zf:
            names = zf.namelist()
            if "data.json" not in names:
                return Response(
                    {"detail": "Archive invalide : data.json introuvable."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            data_content = zf.read("data.json")
            try:
                json.loads(data_content)
            except json.JSONDecodeError:
                return Response({"detail": "data.json invalide ou corrompu."}, status=status.HTTP_400_BAD_REQUEST)

            try:
                call_command("flush", interactive=False)

                tmp_path = None
                with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as tmp:
                    tmp.write(data_content)
                    tmp_path = tmp.name
                try:
                    call_command("loaddata", tmp_path)
                finally:
                    os.unlink(tmp_path)

                media_root = str(settings.MEDIA_ROOT)
                if os.path.isdir(media_root):
                    shutil.rmtree(media_root)
                os.makedirs(media_root, exist_ok=True)
                for name in names:
                    if name.startswith("media/") and not name.endswith("/"):
                        rel_path = name[len("media/"):]
                        target_path = os.path.join(media_root, rel_path)
                        os.makedirs(os.path.dirname(target_path), exist_ok=True)
                        with zf.open(name) as src, open(target_path, "wb") as dst:
                            shutil.copyfileobj(src, dst)
            except Exception as exc:
                return Response(
                    {"detail": f"Erreur lors de la restauration : {exc}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

        return Response({"detail": "Backup restauré avec succès."})
